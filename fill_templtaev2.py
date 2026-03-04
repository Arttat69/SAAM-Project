# ============================================================
# TEMPLATE FILLER - PART I SAAM PROJECT
# Fills Template for Part I-SAAM.xlsx with results
#
# Template layout (from instructor):
#   Stats   → B3:B8  = VW values  |  C3:C8  = MV values
#   Monthly → F3:F146 = VW returns | G3:G146 = MV returns
#   (row 3 = Jan-2014, row 146 = Dec-2025)
#   Values must be raw decimals, NOT percentages (per template instructions)
# ============================================================

import os
import pandas as pd
import openpyxl
from datetime import datetime

print("=" * 60)
print("TEMPLATE FILLER FOR PART I")
print("=" * 60)

# Directory where Part I results CSVs are stored
RESULTS_DIR = "resultsPart1"

# ============================================================
# 1. LOAD RESULTS
# ============================================================

print("\n[1] Loading results from CSV files...")

results_path = os.path.join(RESULTS_DIR, "part1_results.csv")
summary_path = os.path.join(RESULTS_DIR, "part1_summary_statistics.csv")

try:
    results = pd.read_csv(results_path)
    results['Date'] = pd.to_datetime(results['Date'])
    results = results.sort_values('Date').reset_index(drop=True)
    print(f"   ✓ Loaded results from '{results_path}': {len(results)} months")
    if len(results) != 144:
        print(f"   ⚠️  Warning: expected 144 months, got {len(results)}")
except FileNotFoundError:
    print(f"   ✗ Error: '{results_path}' not found!")
    print("   Make sure you have run the Part I script and that the resultsPart1 folder exists.")
    exit(1)

try:
    summary = pd.read_csv(summary_path, index_col=0)
    print(f"   ✓ Loaded summary statistics from '{summary_path}'")
except FileNotFoundError:
    print(f"   ✗ Error: '{summary_path}' not found!")
    exit(1)

mv = summary.loc['Minimum Variance']
vw = summary.loc['Value Weighted']

# ============================================================
# 2. LOAD TEMPLATE
# ============================================================

print("\n[2] Loading template...")

try:
    wb = openpyxl.load_workbook('Template for Part I-SAAM.xlsx')
    print(f"   ✓ Template loaded")
    print(f"   Sheets found: {wb.sheetnames}")
except FileNotFoundError:
    print("   ✗ Error: 'Template for Part I-SAAM.xlsx' not found!")
    exit(1)

ws = wb.active
print(f"   Working with sheet: '{ws.title}'")

# ============================================================
# 3. FILL SUMMARY STATISTICS  (B3:C8)
# ============================================================
# Template row order (rows 3–8):
#   3 = Annualized average return
#   4 = Annualized volatility
#   5 = Annualized cumulative return
#   6 = Sharpe ratio
#   7 = Minimum monthly return
#   8 = Maximum monthly return
#
# Col B = Value-weighted | Col C = Minimum variance
# All values as raw decimals (e.g. 0.0765, NOT 7.65%)
# ============================================================

print("\n[3] Filling summary statistics (B3:C8)...")

stats_map = {
    3: ('Annualized Return',    vw['Annualized Return'],    mv['Annualized Return']),
    4: ('Annualized Volatility', vw['Annualized Volatility'], mv['Annualized Volatility']),
    5: ('Cumulative Return',    vw['Cumulative Return'],    mv['Cumulative Return']),
    6: ('Sharpe Ratio',         vw['Sharpe Ratio'],         mv['Sharpe Ratio']),
    7: ('Min Monthly Return',   vw['Min Monthly Return'],   mv['Min Monthly Return']),
    8: ('Max Monthly Return',   vw['Max Monthly Return'],   mv['Max Monthly Return']),
}

for row_idx, (label, vw_val, mv_val) in stats_map.items():
    ws.cell(row_idx, 2, float(vw_val))  # col B = VW
    ws.cell(row_idx, 3, float(mv_val))  # col C = MV
    print(f"   Row {row_idx} ({label}): VW={vw_val:.6f}  MV={mv_val:.6f}")

print("   ✓ Summary statistics filled")

# ============================================================
# 4. FILL MONTHLY RETURNS  (F3:G146)
# ============================================================
# Col F = Value-weighted returns
# Col G = Minimum variance returns
# Row 3 = Jan-2014, Row 146 = Dec-2025
# Values as raw decimals (e.g. 0.0731, NOT 7.31%)
# ============================================================

print("\n[4] Filling monthly returns (F3:G146)...")

for i, row in results.iterrows():
    excel_row = 3 + i   # row 3 = Jan-2014
    ws.cell(excel_row, 6, float(row['VW_Return']))  # col F = VW
    ws.cell(excel_row, 7, float(row['MV_Return']))  # col G = MV

last_row = 3 + len(results) - 1
print(f"   ✓ Filled rows 3 to {last_row} ({len(results)} months)")
print(f"   First: {results['Date'].iloc[0].strftime('%b-%Y')} → "
      f"VW={results['VW_Return'].iloc[0]:.4f}, MV={results['MV_Return'].iloc[0]:.4f}")
print(f"   Last:  {results['Date'].iloc[-1].strftime('%b-%Y')} → "
      f"VW={results['VW_Return'].iloc[-1]:.4f}, MV={results['MV_Return'].iloc[-1]:.4f}")

# ============================================================
# 5. SAVE
# ============================================================

print("\n[5] Saving filled template...")

output_filename = 'Template-Part-I-FILLED.xlsx'
wb.save(output_filename)
wb.close()

print(f"   ✓ Saved: {output_filename}")

# ============================================================
# SUMMARY
# ============================================================

print("\n" + "=" * 60)
print("TEMPLATE FILLING COMPLETE")
print("=" * 60)

print(f"\n✅ Output file: {output_filename}")
print(f"\n📊 Data filled:")
print(f"   • Stats (B3:C8):       6 rows × 2 portfolios")
print(f"   • Monthly (F3:G{last_row}): {len(results)} months × 2 portfolios")
print(f"\n📈 Key results:")
print(f"   MV  Ann. Return  = {mv['Annualized Return']:.4f}  | VW = {vw['Annualized Return']:.4f}")
print(f"   MV  Ann. Vol     = {mv['Annualized Volatility']:.4f}  | VW = {vw['Annualized Volatility']:.4f}")
print(f"   MV  Sharpe       = {mv['Sharpe Ratio']:.4f}  | VW = {vw['Sharpe Ratio']:.4f}")
print(f"   MV  Cum. Return  = {mv['Cumulative Return']:.4f} (+{mv['Cumulative Return']*100:.1f}%) | "
      f"VW = {vw['Cumulative Return']:.4f} (+{vw['Cumulative Return']*100:.1f}%)")

print("\n" + "=" * 60)
print("READY FOR SUBMISSION")
print("=" * 60)
print("\nNext steps:")
print("1. Open Template-Part-I-FILLED.xlsx")
print("2. Review stats in cells B3:C8")
print("3. Review monthly returns in columns F and G")
print("4. Insert your cumulative return plot in row 9")
print("5. Submit by April 12, 2026 at midnight")
