import os
import re
import warnings
from datetime import datetime

try:
    import cvxpy as cp
except ImportError:
    cp = None
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from scipy.optimize import minimize

warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────
REGION_CODE         = "PAC"
LOW_PRICE_THRESHOLD = 0.5
WINDOW_YEARS        = 10
MIN_OBS_MONTHS      = 36
STALE_THRESHOLD     = 0.50
START_YEAR_OOS      = 2014
END_YEAR_OOS        = 2025
DECISION_YEARS      = list(range(2013, 2025))
Y0                  = 2013
THETA               = 0.10
RIDGE               = 1e-8
SLSQP_MAXITER       = 400
SLSQP_FTOL          = 1e-9
USE_LEDOIT_WOLF     = False

RESULTS_PART1       = "resultsPart1"
RESULTS_PART2       = "ResultsPart2_FINAL"

CO2_FILE            = "DS_CO2_SCOPE_1_Y_2025.xlsx"
RF_FILE             = "Risk_Free_Rate_2025.xlsx"

try:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
except NameError:
    BASE_DIR = os.getcwd()
DATA_DIR = os.path.join(BASE_DIR, "data")

os.makedirs(RESULTS_PART1, exist_ok=True)
os.makedirs(RESULTS_PART2, exist_ok=True)


def data_path(filename: str) -> str:
    candidate = os.path.join(DATA_DIR, filename)
    if os.path.exists(candidate):
        return candidate
    candidate = os.path.join(BASE_DIR, filename)
    if os.path.exists(candidate):
        return candidate
    raise FileNotFoundError(f"Could not find required file: {filename}")


# ─────────────────────────────────────────────────────────────────────────────
# SHARED HELPERS — loading & cleaning
# ─────────────────────────────────────────────────────────────────────────────

def load_datastream_wide(filepath, sheet=0):
    df = pd.read_excel(filepath, sheet_name=sheet, header=0, engine="openpyxl")
    df.columns = ["NAME", "ISIN"] + list(df.columns[2:])
    df = df[~df["NAME"].astype(str).str.startswith("$$ER", na=False)]
    df = df.dropna(subset=["ISIN"])
    df["ISIN"] = df["ISIN"].astype(str).str.strip()
    df["NAME"] = df["NAME"].astype(str).str.strip()
    df = df[df["ISIN"] != ""].set_index("ISIN")
    return df


def load_annual_panel(filepath: str) -> pd.DataFrame:
    df = pd.read_excel(filepath, engine="openpyxl")
    df.columns = ["NAME", "ISIN"] + list(df.columns[2:])
    df = df[~df["NAME"].astype(str).str.startswith("$$ER", na=False)]
    df = df.dropna(subset=["ISIN"])
    df["ISIN"] = df["ISIN"].astype(str).str.strip()
    df = df[df["ISIN"] != ""].set_index("ISIN").drop(columns=["NAME"], errors="ignore")
    df.columns = df.columns.astype(int)
    df = df.apply(pd.to_numeric, errors="coerce")
    return df


def load_rf_monthly(filepath):
    df = pd.read_excel(filepath, engine="openpyxl", index_col=0)
    df.index = df.index.astype(str).str.strip()
    dates = pd.to_datetime(df.index, format="%Y%m") + pd.offsets.MonthEnd(0)
    rf = pd.to_numeric(df.iloc[:, 0], errors="coerce") / 100.0
    rf.index = dates
    return rf.dropna().sort_index()


def parse_monthly_columns(cols):
    parsed, keep = [], []
    for c in cols:
        if c == "NAME":
            continue
        dt = pd.to_datetime(str(c), errors="coerce", dayfirst=True)
        if pd.notna(dt):
            keep.append(c)
            parsed.append(pd.Timestamp(dt).normalize())
    return keep, parsed


def extract_delist_date(name_str):
    m = re.search(r"(?:DELIST|DEAD)\.(\d{2}/\d{2}/\d{2,4})", str(name_str))
    if not m:
        return None
    s = m.group(1)
    for fmt in ("%d/%m/%y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None


def forward_fill_middle_only(price_df, dates):
    arr = price_df.to_numpy(dtype=float, copy=True)
    for i in range(arr.shape[0]):
        row = arr[i, :]
        valid_idx = np.where(~np.isnan(row))[0]
        if valid_idx.size == 0:
            continue
        first, last = valid_idx[0], valid_idx[-1]
        last_val = row[first]
        for k in range(first + 1, last + 1):
            if np.isnan(row[k]):
                row[k] = last_val
            else:
                last_val = row[k]
        arr[i, :] = row
    return pd.DataFrame(arr, index=price_df.index, columns=dates)


def clean_monthly_ri_prices(raw_ri_m, dates, low_price_threshold=0.5,
                            preserve_december_missing=True):
    prices = raw_ri_m.copy().apply(pd.to_numeric, errors="coerce")
    prices = prices.mask(prices < low_price_threshold)
    dec_cols = [d for d in dates if d.month == 12]
    orig_dec_missing = prices[dec_cols].isna() if preserve_december_missing else None
    prices_filled = forward_fill_middle_only(prices, dates)
    if preserve_december_missing and len(dec_cols) > 0:
        prices_filled.loc[:, dec_cols] = prices_filled.loc[:, dec_cols].mask(orig_dec_missing)
    return prices_filled


def apply_delisting_to_returns(returns_df, price_df, delist_dates, dates):
    out = returns_df.copy()
    date_index = pd.Index(dates)
    for isin, ddate in delist_dates.items():
        if ddate is None or isin not in out.index or isin not in price_df.index:
            continue
        dts = pd.Timestamp(ddate).normalize()
        price_series = price_df.loc[isin]
        on_after_delist = date_index[date_index >= dts]
        if len(on_after_delist) == 0:
            continue
        default_dcol = on_after_delist.min()
        future_window = date_index[date_index >= default_dcol]
        future_prices = price_series.loc[future_window]
        if future_prices.notna().any():
            if not (pd.isna(price_series.loc[default_dcol]) and future_prices.isna().all()):
                continue
        prior_valid = date_index[(date_index < default_dcol) & price_series.loc[date_index].notna()]
        if len(prior_valid) > 0:
            last_valid = prior_valid.max()
            candidate_window = date_index[(date_index > last_valid) & (date_index <= default_dcol)]
            missing = candidate_window[price_series.loc[candidate_window].isna()]
            dcol = missing.min() if len(missing) > 0 else default_dcol
        else:
            dcol = default_dcol
        out.at[isin, dcol] = -1.0
        after = date_index[date_index > dcol]
        if len(after) > 0:
            out.loc[isin, after] = np.nan
    return out


def apply_terminal_loss_for_permanent_disappearances(returns_df, price_df, dates):
    out = returns_df.copy()
    date_index = pd.Index(dates)
    for isin in out.index:
        if isin not in price_df.index:
            continue
        price_series = price_df.loc[isin]
        valid_dates = date_index[price_series.loc[date_index].notna()]
        if len(valid_dates) == 0:
            continue
        last_valid_price_date = valid_dates.max()
        tail_dates = date_index[date_index > last_valid_price_date]
        if len(tail_dates) == 0:
            continue
        if price_series.loc[tail_dates].isna().all():
            first_missing = tail_dates.min()
            if pd.isna(out.at[isin, first_missing]):
                out.at[isin, first_missing] = -1.0
            later = date_index[date_index > first_missing]
            if len(later) > 0:
                out.loc[isin, later] = np.nan
    return out


def adjust_mv_caps_for_terminal_events(mv_caps, ri_returns, dates):
    caps = mv_caps.copy()
    date_index = pd.Index(dates)
    for isin in caps.index.intersection(ri_returns.index):
        series = ri_returns.loc[isin, date_index]
        minus100_dates = date_index[series == -1.0]
        if len(minus100_dates) == 0:
            continue
        for d in minus100_dates:
            after = date_index[date_index > d]
            if len(after) == 0 or series.loc[after].isna().all():
                caps.loc[isin, date_index[date_index >= d]] = np.nan
                break
    return caps


# ─────────────────────────────────────────────────────────────────────────────
# SHARED HELPERS — investment set, moments, optimisation, performance
# ─────────────────────────────────────────────────────────────────────────────

def year_end_col(dates, year):
    decs = [d for d in dates if d.year == year and d.month == 12]
    return max(decs) if decs else None


def window_cols(dates, year_end, window_years=10):
    start = pd.Timestamp(f"{year_end - window_years + 1}-01-01")
    end = pd.Timestamp(f"{year_end}-12-31")
    return [d for d in dates if start <= d <= end]


def stale_mask(returns_df, cols, threshold=0.50):
    sub = returns_df[cols]
    denom = sub.notna().sum(axis=1).replace(0, np.nan)
    frac0 = (sub == 0).sum(axis=1) / denom
    return frac0 > threshold


def build_investment_set(ri_prices, ri_returns, dates, year_end, co2_panel=None):
    dec_col = year_end_col(dates, year_end)
    if dec_col is None:
        raise ValueError(f"No December column for year {year_end}.")
    cols = window_cols(dates, year_end, WINDOW_YEARS)
    ok_price = ri_prices[dec_col].notna()
    n_obs = ri_returns[cols].notna().sum(axis=1)
    ok_obs = n_obs >= MIN_OBS_MONTHS
    ok_stale = ~stale_mask(ri_returns, cols, STALE_THRESHOLD).fillna(True)
    ok_carbon = pd.Series(True, index=ri_prices.index)
    if co2_panel is not None:
        if year_end in co2_panel.columns:
            ok_carbon = co2_panel[year_end].reindex(ri_prices.index).notna().fillna(False)
        else:
            print(f"Warning: year {year_end} not in CO2 panel — carbon filter skipped.")
    eligible = ri_prices.index[ok_price & ok_obs & ok_stale & ok_carbon]
    return list(eligible), cols

# Methodological note:
# Expected returns are estimated from the trailing 10-year monthly window.
# The covariance matrix is implemented as a custom missing-data approximation:
# asset means are computed using available monthly observations, missing
# demeaned returns are set to zero, and the matrix is scaled by the minimum
# number of valid observations across assets. A very small diagonal ridge is
# added only for numerical stability. This is not a textbook pairwise sample
# covariance estimator and is disclosed explicitly in the report.
def estimate_moments(ri_returns, isins, cols, ridge=1e-8):
   

    R = ri_returns.loc[isins, cols].astype(float)
    mu = R.mean(axis=1, skipna=True).to_numpy()
    R_np = R.to_numpy()
    mu_col = np.nanmean(R_np, axis=1, keepdims=True)
    demeaned = np.where(np.isnan(R_np), 0.0, R_np - mu_col)
    valid_counts = np.isfinite(R_np).sum(axis=1)
    tau = valid_counts.min()
    Sigma = demeaned @ demeaned.T / tau
    if np.isnan(Sigma).any():
        raise ValueError("Covariance matrix contains NaN values after estimation.")
    Sigma = Sigma + ridge * np.eye(Sigma.shape[0])
    return mu, Sigma


def solve_min_variance(Sigma):
    n = Sigma.shape[0]
    a0 = np.full(n, 1.0 / n)
    res = minimize(
        lambda a: float(a @ Sigma @ a),
        a0,
        method="SLSQP",
        bounds=[(0.0, 1.0)] * n,
        constraints=[{"type": "eq", "fun": lambda a: np.sum(a) - 1.0}],
        options={"maxiter": SLSQP_MAXITER, "ftol": SLSQP_FTOL, "disp": False},
    )
    if not res.success:
        raise RuntimeError(f"SLSQP failed: {res.message}")
    w = np.where(res.x < 1e-10, 0.0, res.x)
    return w / w.sum()

# Methodological note:
# The tracking-error portfolios in Sections 3.3 and 4.1 are constructed relative
# to the value-weighted benchmark on the same annual eligible universe. The
# objective minimizes quadratic tracking error, i.e. (w - w_vw)' Σ (w - w_vw),
# which is equivalent to minimizing squared ex-ante tracking error.
def solve_cvxpy(Sigma_arr, e_c, cf_target, w_ref=None, mode="mv"):
    N = Sigma_arr.shape[0]

    def _solve_slsqp():
        x0 = w_ref.copy() if (mode == "te" and w_ref is not None) else np.full(N, 1.0 / N)
        def objective(a):
            if mode == "mv":
                diff = a
            else:
                diff = a - w_ref
            return float(diff @ Sigma_arr @ diff)
        constraints = [
            {"type": "eq", "fun": lambda a: np.sum(a) - 1.0},
            {"type": "ineq", "fun": lambda a: cf_target - float(a @ e_c)},
        ]
        res = minimize(
            objective,
            x0,
            method="SLSQP",
            bounds=[(0.0, 1.0)] * N,
            constraints=constraints,
            options={"maxiter": 1000, "ftol": 1e-9, "disp": False},
        )
        if not res.success:
            return None
        w = np.maximum(res.x, 0.0)
        s = w.sum()
        return w / s if s > 0 else None

    if cp is None:
        return _solve_slsqp()

    Sp = cp.psd_wrap(Sigma_arr)
    alpha = cp.Variable(N)
    if mode == "mv":
        obj = cp.Minimize(cp.quad_form(alpha, Sp))
    else:
        obj = cp.Minimize(cp.quad_form(alpha - w_ref, Sp))
    prob = cp.Problem(obj, [cp.sum(alpha) == 1, alpha >= 0, alpha @ e_c <= cf_target])
    for solver in (cp.OSQP, cp.SCS):
        try:
            kw = {"eps_abs": 1e-8, "eps_rel": 1e-8, "max_iter": 20_000} if solver == cp.OSQP else {}
            prob.solve(solver=solver, **kw)
        except Exception:
            continue
        if alpha.value is not None and prob.status in ("optimal", "optimal_inaccurate"):
            w = np.maximum(alpha.value, 0.0)
            s = w.sum()
            return w / s if s > 0 else np.ones(N) / N
    return _solve_slsqp()

# Methodological note:
# Realized monthly returns are not silently replaced with zero for invested
# positions. Missing returns are only acceptable when the associated portfolio
# weight is already economically zero. Otherwise the code raises an error.
# This preserves the integrity of the out-of-sample portfolio return calculation.
def compute_mv_oos_returns(portfolios, ri_returns, dates):
    out_r, out_d = [], []
    WEIGHT_TOL = 1e-12
    for year_end, info in portfolios.items():
        invest_year = year_end + 1
        year_months = [d for d in dates if pd.Timestamp(f"{invest_year}-01-01") <= d <= pd.Timestamp(f"{invest_year}-12-31")]
        if not year_months:
            continue
        isins = info["isins"]
        w = info["weights"].copy()
        for d in year_months:
            r_series = ri_returns.loc[isins, d].copy()
            missing_mask = r_series.isna().to_numpy()
            bad_mask = missing_mask & (w > WEIGHT_TOL)
            if bad_mask.any():
                bad_isins = list(np.array(isins)[bad_mask])
                bad_weights = w[bad_mask]
                raise ValueError(
                    f"Missing realized returns for invested names on {d.date()}. ISINs={bad_isins}, weights={bad_weights.tolist()}"
                )
            r_i = r_series.fillna(0.0).to_numpy()
            r_p = float(w @ r_i)
            out_r.append(r_p)
            out_d.append(d)
            denom = 1.0 + r_p
            if denom <= 0:
                raise ValueError(f"Portfolio value collapsed to non-positive level on {d.date()} (1 + r_p = {denom}).")
            w = w * (1.0 + r_i) / denom
            w = np.where(w < WEIGHT_TOL, 0.0, w)
            if w.sum() <= 0:
                raise ValueError(f"Portfolio weights sum to zero after update on {d.date()}.")
            w = w / w.sum()
    return pd.Series(out_r, index=pd.to_datetime(out_d)).sort_index()


def compute_vw_oos_returns(portfolios, ri_returns, mv_caps, dates, start_year=2014, end_year=2025):
    out_r, out_d = [], []
    dates_sorted = sorted(dates)
    pos = {d: i for i, d in enumerate(dates_sorted)}
    for year_end, info in portfolios.items():
        invest_year = year_end + 1
        if invest_year < start_year or invest_year > end_year:
            continue
        eligible_isins = list(info["isins"])
        year_months = [d for d in dates_sorted if pd.Timestamp(f"{invest_year}-01-01") <= d <= pd.Timestamp(f"{invest_year}-12-31")]
        for d in year_months:
            i = pos.get(d)
            if i is None or i == 0:
                continue
            d_prev = dates_sorted[i - 1]
            if d_prev not in mv_caps.columns:
                raise ValueError(f"Lagged market-cap column missing for benchmark month {d_prev.date()}")
            if d not in ri_returns.columns:
                raise ValueError(f"Return column missing for benchmark month {d.date()}")
            caps = mv_caps.reindex(eligible_isins)[d_prev]
            rets = ri_returns.reindex(eligible_isins)[d]
            caps_missing = caps.isna()
            rets_missing = rets.isna()
            inconsistent = caps_missing ^ rets_missing
            if inconsistent.any():
                bad_isins = list(caps.index[inconsistent])
                raise ValueError(
                    f"VW benchmark inconsistent data on {d.date()}.\nDropped ISINs: {bad_isins}\nReason: market cap and return availability do not match."
                )
            valid = (~caps_missing) & (~rets_missing)
            if valid.sum() == 0:
                raise ValueError(f"No valid benchmark constituents on {d.date()}")
            cap_sum = caps.loc[valid].sum()
            if pd.isna(cap_sum) or cap_sum <= 0:
                raise ValueError(f"Invalid lagged market-cap sum for benchmark on {d.date()}")
            w = caps.loc[valid] / cap_sum
            r_p = float(w.to_numpy() @ rets.loc[valid].to_numpy())
            out_r.append(r_p)
            out_d.append(d)
    return pd.Series(out_r, index=pd.to_datetime(out_d)).sort_index()


def compute_oos_returns(portfolios_dict, ri_ret, dates, start_yr, end_yr):
    out_r, out_d = [], []
    WEIGHT_TOL = 1e-12

    for year_end in sorted(portfolios_dict.keys()):
        invest_year = year_end + 1
        if invest_year < start_yr or invest_year > end_yr:
            continue

        year_months = [d for d in dates if d.year == invest_year]
        if not year_months:
            continue

        isins = portfolios_dict[year_end]["isins"]
        w = portfolios_dict[year_end]["weights"].copy()

        for d in year_months:
            r_series = ri_ret.loc[isins, d].copy()

            missing_mask = r_series.isna().to_numpy()
            bad_mask = missing_mask & (w > WEIGHT_TOL)

            if bad_mask.any():
                bad_isins = list(np.array(isins)[bad_mask])
                bad_weights = w[bad_mask]
                raise ValueError(
                    f"Missing realized returns for invested names on {d.date()}. "
                    f"ISINs={bad_isins}, weights={bad_weights.tolist()}"
                )

            r_i = r_series.fillna(0.0).to_numpy()

            r_p = float(w @ r_i)
            out_r.append(r_p)
            out_d.append(d)

            denom = 1.0 + r_p
            if denom <= 0:
                raise ValueError(
                    f"Portfolio value collapsed to non-positive level on {d.date()} "
                    f"(1 + r_p = {denom})."
                )

            w = w * (1.0 + r_i) / denom
            w = np.where(w < WEIGHT_TOL, 0.0, w)

            if w.sum() <= 0:
                raise ValueError(f"Portfolio weights sum to zero after update on {d.date()}.")

            w = w / w.sum()

    return pd.Series(out_r, index=pd.to_datetime(out_d)).sort_index()


def validate_oos_series(series: pd.Series, name: str):
    expected_months = pd.date_range(start=f"{START_YEAR_OOS}-01-31", end=f"{END_YEAR_OOS}-12-31", freq="M")
    s = series.copy()
    s.index = pd.to_datetime(s.index) + pd.offsets.MonthEnd(0)
    s = s[~s.index.duplicated(keep="last")].reindex(expected_months)
    if s.isna().any():
        missing = [d.strftime("%Y-%m-%d") for d in s.index[s.isna()]]
        raise ValueError(f"{name} is missing OOS returns for months: {missing}")
    if len(s) != 144:
        raise ValueError(f"{name} has {len(s)} observations, expected 144.")
    return s


def perf_stats(r, rf=None):
    r = r.dropna()
    T = len(r)
    if T == 0:
        return {
            "Annualized Average Return": np.nan,
            "Annualized Volatility": np.nan,
            "Annualized Cumulative Return": np.nan,
            "Sharpe Ratio": np.nan,
            "Minimum Monthly Return": np.nan,
            "Maximum Monthly Return": np.nan,
        }
    rf_aligned = rf.reindex(r.index).fillna(0.0) if rf is not None else pd.Series(0.0, index=r.index)
    excess = r - rf_aligned
    mean_m = r.mean()
    vol_m = r.std(ddof=1)
    vol_excess = excess.std(ddof=1)
    return {
        "Annualized Average Return": 12.0 * mean_m,
        "Annualized Volatility": vol_m * np.sqrt(12.0),
        "Annualized Cumulative Return": (1.0 + r).prod() ** (12.0 / T) - 1.0,
        "Sharpe Ratio": (excess.mean() / vol_excess) * np.sqrt(12.0) if vol_excess > 0 else np.nan,
        "Minimum Monthly Return": r.min(),
        "Maximum Monthly Return": r.max(),
    }


def perf_stats_extended(r: pd.Series, rf: pd.Series = None) -> dict:
    r = r.dropna()
    T = len(r)
    keys = [
        "Annualized Average Return", "Annualized Volatility",
        "Annualized Cumulative Return", "Sharpe Ratio",
        "Minimum Monthly Return", "Maximum Monthly Return",
        "VaR_95 (monthly)", "VaR_99 (monthly)",
        "ES_95 (monthly)", "ES_99 (monthly)", "Max_Drawdown",
    ]
    if T == 0:
        return {k: np.nan for k in keys}
    rf_aligned = rf.reindex(r.index).fillna(0.0) if rf is not None else pd.Series(0.0, index=r.index)
    excess = r - rf_aligned
    mean_m = r.mean()
    vol_m = r.std(ddof=1)
    vol_exc = excess.std(ddof=1)
    var95 = -float(np.percentile(r, 5))
    var99 = -float(np.percentile(r, 1))
    tail95 = r[r <= -var95]
    es95 = float(-tail95.mean()) if len(tail95) > 0 else var95
    tail99 = r[r <= -var99]
    es99 = float(-tail99.mean()) if len(tail99) > 0 else var99
    cum = (1.0 + r).cumprod()
    running_max = cum.cummax()
    mdd = float((cum / running_max - 1.0).min())
    return {
        "Annualized Average Return": 12.0 * mean_m,
        "Annualized Volatility": vol_m * np.sqrt(12.0),
        "Annualized Cumulative Return": (1.0 + r).prod() ** (12.0 / T) - 1.0,
        "Sharpe Ratio": (excess.mean() / vol_exc * np.sqrt(12.0) if vol_exc > 0 else np.nan),
        "Minimum Monthly Return": r.min(),
        "Maximum Monthly Return": r.max(),
        "VaR_95 (monthly)": var95,
        "VaR_99 (monthly)": var99,
        "ES_95 (monthly)": es95,
        "ES_99 (monthly)": es99,
        "Max_Drawdown": mdd,
    }


def perf_stats_relative(r_portfolio: pd.Series, r_benchmark: pd.Series, portfolio_weights: dict, benchmark_weights: dict, rf: pd.Series = None) -> dict:
    r_p = r_portfolio.dropna()
    r_b = r_benchmark.reindex(r_p.index).dropna()
    common = r_p.index.intersection(r_b.index)
    r_p, r_b = r_p[common], r_b[common]
    active = r_p - r_b
    ann_active = 12.0 * active.mean()
    te_ann = active.std(ddof=1) * np.sqrt(12.0)
    ir = ann_active / te_ann if te_ann > 0 else np.nan
    active_shares = []
    for Y in sorted(set(portfolio_weights) & set(benchmark_weights)):
        p_info = portfolio_weights[Y]
        b_info = benchmark_weights[Y]
        all_isins = list(set(p_info["isins"]) | set(b_info["isins"]))
        wp = pd.Series(dict(zip(p_info["isins"], p_info["weights"]))).reindex(all_isins).fillna(0.0)
        wb = pd.Series(dict(zip(b_info["isins"], b_info["weights"]))).reindex(all_isins).fillna(0.0)
        active_shares.append(0.5 * np.abs(wp.values - wb.values).sum())
    return {
        "Active Return (ann.)": ann_active,
        "Tracking Error (ann.)": te_ann,
        "Information Ratio": ir,
        "Avg Active Share": float(np.mean(active_shares)) if active_shares else np.nan,
    }


def export_part1_excel_template(template_path, output_path, stats, out_df):
    wb = load_workbook(template_path)
    ws = wb["Sheet1"]
    ws["B3"] = float(stats.loc["Value Weighted", "Annualized Average Return"])
    ws["B4"] = float(stats.loc["Value Weighted", "Annualized Volatility"])
    ws["B5"] = float(stats.loc["Value Weighted", "Annualized Cumulative Return"])
    ws["B6"] = float(stats.loc["Value Weighted", "Sharpe Ratio"])
    ws["B7"] = float(stats.loc["Value Weighted", "Minimum Monthly Return"])
    ws["B8"] = float(stats.loc["Value Weighted", "Maximum Monthly Return"])
    ws["C3"] = float(stats.loc["Minimum Variance", "Annualized Average Return"])
    ws["C4"] = float(stats.loc["Minimum Variance", "Annualized Volatility"])
    ws["C5"] = float(stats.loc["Minimum Variance", "Annualized Cumulative Return"])
    ws["C6"] = float(stats.loc["Minimum Variance", "Sharpe Ratio"])
    ws["C7"] = float(stats.loc["Minimum Variance", "Minimum Monthly Return"])
    ws["C8"] = float(stats.loc["Minimum Variance", "Maximum Monthly Return"])
    if len(out_df) != 144:
        raise ValueError(f"Expected 144 monthly observations, got {len(out_df)}")
    start_row = 3
    for i, row in enumerate(out_df.itertuples(index=False), start=start_row):
        ws[f"F{i}"] = float(row.VW_Return)
        ws[f"G{i}"] = float(row.MV_Return)
    plot_path = os.path.join(os.path.dirname(output_path), "part1_cumulative_plot.png")
    plt.figure(figsize=(9, 4.8))
    plt.plot(out_df["Date"], out_df["VW_CumReturn"], label="Value-weighted portfolio")
    plt.plot(out_df["Date"], out_df["MV_CumReturn"], label="Minimum-variance portfolio")
    plt.title("Cumulative Returns (2014–2025)")
    plt.xlabel("Date")
    plt.ylabel("Growth of $1")
    plt.legend()
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(plot_path, dpi=200, bbox_inches="tight")
    plt.close()
    img = XLImage(plot_path)
    img.width = 520
    img.height = 280
    ws.add_image(img, "B10")
    wb.save(output_path)


# ─────────────────────────────────────────────────────────────────────────────
# SHARED HELPERS — carbon metrics, plots, shrinkage
# ─────────────────────────────────────────────────────────────────────────────

def carbon_intensity_vec(e, r):
    with np.errstate(invalid="ignore", divide="ignore"):
        return np.where(r > 0, e / r, np.nan)

# Methodological note:
# Carbon intensity is undefined when firm-year revenue is missing or non-positive.
# In those cases, WACI is computed over the subset of firms with valid carbon
# intensity data, with portfolio weights re-normalized over that valid subset.
# This choice is made to preserve the same investment universe as Part I and is
# disclosed explicitly in the report as a practical implementation choice.
def waci_metric(w, ci):
    valid = ~np.isnan(ci)
    if not valid.any():
        return np.nan
    w2 = np.where(valid, w, 0.0)
    s = w2.sum()
    return float((w2 / s) @ np.where(valid, ci, 0.0)) if s > 0 else np.nan


def cf_metric(w, e, c):
    with np.errstate(invalid="ignore", divide="ignore"):
        ratio = np.where(c > 0, e / c, 0.0)
    return float(w @ ratio)


def cf_vw_metric(e, c):
    total = c.sum()
    return float(e.sum() / total) if total > 0 else np.nan


def e_over_c_vec(e, c):
    with np.errstate(invalid="ignore", divide="ignore"):
        return np.where(c > 0, e / c, 0.0)


CF_TOL = 1e-4

def validate_cf_constraint(w, e, c, cf_target, label, Y):
    achieved = cf_metric(w, e, c)
    if not np.isfinite(achieved):
        raise ValueError(f"{label} | Y={Y}: achieved CF is not finite.")
    if achieved > cf_target + CF_TOL:
        raise ValueError(f"{label} | Y={Y}: CF constraint violated. Achieved={achieved:.8f}, Target={cf_target:.8f}")
    return achieved


def save_cumret_plot(series_dict, filename, title):
    plt.figure(figsize=(10, 6))
    for label, s in series_dict.items():
        s = s.dropna().sort_index()
        wealth = (1.0 + s).cumprod()
        plt.plot(wealth.index, wealth.values, label=label)
    plt.title(title)
    plt.xlabel("Date")
    plt.ylabel("Cumulative wealth")
    plt.legend()
    plt.tight_layout()
    plt.savefig(filename, dpi=200)
    plt.close()


def save_annual_line_plot(df, filename, title, ylabel):
    plt.figure(figsize=(10, 6))
    for col in df.columns:
        plt.plot(df.index, df[col], marker="o", label=col)
    plt.title(title)
    plt.xlabel("Year")
    plt.ylabel(ylabel)
    plt.legend()
    plt.tight_layout()
    plt.savefig(filename, dpi=200)
    plt.close()

def export_weight_comparison(reference_portfolios, candidate_portfolios, output_csv):
    """
    Compare two sets of annual portfolio weights and export the differences.
    """
    rows = []

    common_years = sorted(set(reference_portfolios.keys()) & set(candidate_portfolios.keys()))

    for Y in common_years:
        ref_info = reference_portfolios[Y]
        cand_info = candidate_portfolios[Y]

        ref_w = pd.Series(ref_info["weights"], index=ref_info["isins"], name="ReferenceWeight")
        cand_w = pd.Series(cand_info["weights"], index=cand_info["isins"], name="CandidateWeight")

        all_isins = sorted(set(ref_w.index) | set(cand_w.index))

        comp = pd.DataFrame(index=all_isins)
        comp["ReferenceWeight"] = ref_w.reindex(all_isins).fillna(0.0)
        comp["CandidateWeight"] = cand_w.reindex(all_isins).fillna(0.0)
        comp["DeltaWeight"] = comp["CandidateWeight"] - comp["ReferenceWeight"]
        comp["AbsDeltaWeight"] = comp["DeltaWeight"].abs()

        conditions = [
            (comp["ReferenceWeight"] == 0) & (comp["CandidateWeight"] > 0),
            (comp["ReferenceWeight"] > 0) & (comp["CandidateWeight"] == 0),
            comp["DeltaWeight"] > 0,
            comp["DeltaWeight"] < 0,
        ]
        labels = ["Added", "Removed", "Overweighted", "Underweighted"]
        comp["ChangeType"] = np.select(conditions, labels, default="Unchanged")

        comp["Year"] = Y + 1
        comp["ISIN"] = comp.index

        rows.append(comp.reset_index(drop=True))

    if rows:
        out = pd.concat(rows, ignore_index=True)
        out = out.sort_values(["Year", "AbsDeltaWeight"], ascending=[True, False])
        out.to_csv(output_csv, index=False)
        print(f"Weight comparison exported: {output_csv}")
    else:
        print(f"Warning: no common years found for {output_csv}")


def export_top_weight_changes(comparison_csv, output_csv, top_n=20):
    """
    From a full weight-comparison CSV, keep only the largest absolute weight changes
    for each year and export them to a separate file.
    """
    df = pd.read_csv(comparison_csv)

    top_df = (
        df.sort_values(["Year", "AbsDeltaWeight"], ascending=[True, False])
          .groupby("Year", group_keys=False)
          .head(top_n)
          .copy()
    )

    top_df.to_csv(output_csv, index=False)
    print(f"Top weight changes exported: {output_csv}")




def ledoit_wolf_cc(X: np.ndarray):
    T, N = X.shape
    S = X.T @ X / T
    var = np.diag(S)
    std = np.sqrt(np.maximum(var, 1e-16))
    std_outer = np.outer(std, std)
    corr = S / std_outer
    np.fill_diagonal(corr, 0.0)
    r_bar = corr.sum() / (N * (N - 1))
    F = r_bar * std_outer
    np.fill_diagonal(F, var)
    X2 = X ** 2
    pi_mat = X2.T @ X2 / T - S ** 2
    pi_hat = pi_mat.sum()
    rho_hat = np.diag(pi_mat).sum()
    X3 = X ** 3
    Theta_II = X3.T @ X / T - S * var[:, None]
    Theta_JJ = X.T @ X3 / T - S * var[None, :]
    ratio_JI = std[None, :] / std[:, None]
    ratio_IJ = std[:, None] / std[None, :]
    rho_mat = (r_bar / 2) * (ratio_JI * Theta_II + ratio_IJ * Theta_JJ)
    rho_hat += rho_mat.sum() - np.diag(rho_mat).sum()
    gamma_hat = np.sum((F - S) ** 2)
    kappa = (pi_hat - rho_hat) / gamma_hat if gamma_hat > 1e-30 else 0.0
    delta = float(np.clip(kappa / T, 0.0, 1.0))
    return delta * F + (1.0 - delta) * S, delta, r_bar


def annual_portfolio_values(monthly_ret: pd.Series, v0: float = 1.0) -> pd.Series:
    values = {Y0: v0}
    v = v0
    for year in range(START_YEAR_OOS, END_YEAR_OOS + 1):
        yr_rets = monthly_ret[monthly_ret.index.year == year]
        if len(yr_rets) > 0:
            v = v * (1.0 + yr_rets).prod()
        values[year] = v
    return pd.Series(values).sort_index()


# ─────────────────────────────────────────────────────────────────────────────
# LOAD COMMON DATA ONCE
# ─────────────────────────────────────────────────────────────────────────────
print("=" * 60)
print("SAAM Project 2026 — Combined Part I + Part II")
print(f"Region: {REGION_CODE} | Scope: Scope 1")
print("=" * 60)

static = pd.read_excel(data_path("Static_2025.xlsx"), engine="openpyxl")
static.columns = ["ISIN", "NAME", "Country", "Region"]
static["ISIN"] = static["ISIN"].astype(str).str.strip()
pac = static[static["Region"] == REGION_CODE].copy().set_index("ISIN")
pac_isins = set(pac.index)
print(f"Pacific firms in Static: {len(pac_isins)}")

ri_y = load_datastream_wide(data_path("DS_RI_T_USD_Y_2025.xlsx"))
ri_y = ri_y[ri_y.index.isin(pac_isins)].copy()
delist_dates = {isin: extract_delist_date(ri_y.at[isin, "NAME"]) for isin in ri_y.index}
print(f"Delisted Pacific firms: {sum(v is not None for v in delist_dates.values())}")

ri_m_raw = load_datastream_wide(data_path("DS_RI_T_USD_M_2025.xlsx"))
keep_cols, parsed_dates = parse_monthly_columns(list(ri_m_raw.columns))
ri_m = ri_m_raw[["NAME"] + keep_cols].copy()
ri_m.columns = ["NAME"] + parsed_dates
ri_m = ri_m[ri_m.index.isin(pac_isins)].copy()
parsed_dates = [d for d in parsed_dates if d <= pd.Timestamp("2025-12-31")]
ri_m = ri_m[["NAME"] + parsed_dates]
print(f"Monthly RI: {ri_m.shape[0]} firms, {len(parsed_dates)} months")

mv_m_raw = load_datastream_wide(data_path("DS_MV_T_USD_M_2025.xlsx"))
keep_cols_mv, _mv = parse_monthly_columns(list(mv_m_raw.columns))
mv_m = mv_m_raw[["NAME"] + keep_cols_mv].copy()
mv_m.columns = ["NAME"] + _mv
mv_m = mv_m[mv_m.index.isin(pac_isins)].copy()
mv_m = mv_m[["NAME"] + parsed_dates].copy()
mv_m[parsed_dates] = mv_m[parsed_dates].apply(pd.to_numeric, errors="coerce")
print(f"Monthly MV: {mv_m.shape[0]} firms")

common_isins = pac_isins.intersection(set(ri_m.index)).intersection(set(mv_m.index))
ri_m = ri_m.loc[list(common_isins)].copy()
mv_m = mv_m.loc[list(common_isins)].copy()
print(f"Common ISINs: {len(common_isins)}")

ri_prices_universe = clean_monthly_ri_prices(ri_m[parsed_dates], parsed_dates, LOW_PRICE_THRESHOLD, preserve_december_missing=True)
ri_prices_returns = clean_monthly_ri_prices(ri_m[parsed_dates], parsed_dates, LOW_PRICE_THRESHOLD, preserve_december_missing=False)
all_missing = ri_prices_returns.isna().all(axis=1)
if all_missing.any():
    missing_isins = ri_prices_returns.index[all_missing]
    ri_prices_returns = ri_prices_returns.loc[~all_missing].copy()
    ri_prices_universe = ri_prices_universe.loc[~all_missing].copy()
    mv_m = mv_m.drop(index=missing_isins, errors="ignore")
    print(f"Dropped fully-missing RI rows: {all_missing.sum()}")

ri_returns = ri_prices_returns.pct_change(axis=1, fill_method=None)
ri_returns = apply_delisting_to_returns(ri_returns, ri_prices_returns, {k: v for k, v in delist_dates.items() if k in ri_returns.index}, parsed_dates)
ri_returns = apply_terminal_loss_for_permanent_disappearances(ri_returns, ri_prices_returns, parsed_dates)
mv_caps_adjusted = adjust_mv_caps_for_terminal_events(mv_m[parsed_dates], ri_returns, parsed_dates)
print("Monthly returns and adjusted MV caps ready.")

co2_panel = load_annual_panel(data_path(CO2_FILE))
co2_panel = co2_panel.T.ffill().T
rev_raw = load_annual_panel(data_path("DS_REV_Y_2025.xlsx"))
cap_ann_raw = load_annual_panel(data_path("DS_MV_T_USD_Y_2025.xlsx"))
rev_m = rev_raw / 1000.0
co2_ff = co2_panel.T.ffill().T
revM_ff = rev_m.T.ffill().T
capA_ff = cap_ann_raw.T.ffill().T
print(f"CO2 panel: {co2_ff.shape[0]} ISINs, years {co2_ff.columns.min()}–{co2_ff.columns.max()}")

_tmp = pd.read_excel(data_path(CO2_FILE), engine="openpyxl")
_tmp.columns = ["NAME", "ISIN"] + list(_tmp.columns[2:])
_tmp = _tmp[~_tmp["NAME"].astype(str).str.startswith("$$ER", na=False)].dropna(subset=["ISIN"])
_tmp["ISIN"] = _tmp["ISIN"].astype(str).str.strip()
firm_names = _tmp.set_index("ISIN")["NAME"]

try:
    rf_monthly = load_rf_monthly(data_path(RF_FILE))
    print(f"RF loaded: {len(rf_monthly)} obs, mean={rf_monthly.mean() * 100:.3f} %/month")
except Exception as exc:
    print(f"WARNING: could not load RF ({exc}). Using rf=0.")
    rf_monthly = None


def get_carbon_vectors(Y: int, isins: list):
    e = co2_ff[Y].reindex(isins).fillna(0.0).values if Y in co2_ff.columns else np.zeros(len(isins))
    r = revM_ff[Y].reindex(isins).fillna(np.nan).values if Y in revM_ff.columns else np.full(len(isins), np.nan)
    c = capA_ff[Y].reindex(isins).fillna(0.0).values if Y in capA_ff.columns else np.zeros(len(isins))
    return e, r, c


# ─────────────────────────────────────────────────────────────────────────────
# PART I — Standard portfolio allocation
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 60)
print("PART I — Standard Portfolio Allocation")
print("=" * 60)

part1_portfolios = {}
for year_end in range(START_YEAR_OOS - 1, END_YEAR_OOS):
    elig, cols = build_investment_set(ri_prices_universe, ri_returns, parsed_dates, year_end, co2_panel=co2_panel)
    if len(elig) < 2:
        print(f"Year {year_end}: eligible={len(elig)} (skipped)")
        continue
    _, Sigma = estimate_moments(ri_returns, elig, cols)
    w = solve_min_variance(Sigma)
    part1_portfolios[year_end] = {"isins": elig, "weights": w}
    print(f"Year {year_end}: eligible={len(elig)}, max_w={w.max():.4f}")

mv_r = validate_oos_series(compute_mv_oos_returns(part1_portfolios, ri_returns, parsed_dates), "Minimum-variance portfolio")
vw_r = validate_oos_series(compute_vw_oos_returns(part1_portfolios, ri_returns, mv_caps_adjusted, parsed_dates, START_YEAR_OOS, END_YEAR_OOS), "Value-weighted benchmark")

expected_months = pd.date_range(start=f"{START_YEAR_OOS}-01-31", end=f"{END_YEAR_OOS}-12-31", freq="M")
part1_out = pd.DataFrame({"Date": expected_months, "MV_Return": mv_r.values, "VW_Return": vw_r.values})
part1_out["MV_CumReturn"] = (1.0 + part1_out["MV_Return"]).cumprod()
part1_out["VW_CumReturn"] = (1.0 + part1_out["VW_Return"]).cumprod()
part1_out.to_csv(os.path.join(RESULTS_PART1, "part1_results.csv"), index=False)

part1_stats = pd.DataFrame({
    "Minimum Variance": perf_stats(mv_r, rf=rf_monthly),
    "Value Weighted": perf_stats(vw_r, rf=rf_monthly),
}).T
part1_stats.to_csv(os.path.join(RESULTS_PART1, "part1_summary_statistics.csv"))

part1_comp = pd.concat([
    pd.DataFrame({"Year": year_end + 1, "ISIN": info["isins"], "Weight": info["weights"]})
    for year_end, info in part1_portfolios.items()
], ignore_index=True)
part1_comp.to_csv(os.path.join(RESULTS_PART1, "part1_portfolio_compositions.csv"), index=False)

export_part1_excel_template(
    template_path=data_path("Template for Part I-SAAM.xlsx"),
    output_path=os.path.join(RESULTS_PART1, "Template_for_Part_I_SAAM_FILLED.xlsx"),
    stats=part1_stats,
    out_df=part1_out,
)

mv_weights_p1 = {yr - 1: grp.set_index("ISIN")["Weight"] for yr, grp in part1_comp.groupby("Year")}
print("Part I outputs written to:", RESULTS_PART1)


# ─────────────────────────────────────────────────────────────────────────────
# PART II — Carbon objectives and net-zero analysis
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 60)
print("PART II — Portfolio Allocation with Carbon Emission Reduction")
print("=" * 60)

inv_sets_p2 = {}
sigma_p2 = {}
vw_weights_p2 = {}
lw_deltas = {}
lw_r_bars = {}
# Methodological note:
# To preserve strict comparability between Parts I and II, the Part II
# investable universe is set equal to the Part I universe for each decision
# year Y. This avoids mixing the effect of carbon constraints with the effect
# of changing the underlying eligible universe.
for Y in DECISION_YEARS:
    if Y not in part1_portfolios:
        print(f"Y={Y}: not available in Part I portfolios — skipping.")
        continue

    # EXACT SAME universe as Part I
    elig = list(part1_portfolios[Y]["isins"])
    cols = window_cols(parsed_dates, Y, WINDOW_YEARS)

    if len(elig) < 2:
        print(f"Y={Y}: only {len(elig)} eligible firms — skipping.")
        continue

    _, Sigma_sample = estimate_moments(ri_returns, elig, cols)

    if USE_LEDOIT_WOLF:
        R_mat = ri_returns.loc[elig, cols].fillna(0.0).to_numpy().T
        R_dem = R_mat - R_mat.mean(axis=0, keepdims=True)
        Sigma, delta_lw, r_bar_lw = ledoit_wolf_cc(R_dem)
        lw_deltas[Y] = delta_lw
        lw_r_bars[Y] = r_bar_lw
        msg = f", δ̂={delta_lw:.4f}, r̄={r_bar_lw:.4f}"
    else:
        Sigma = Sigma_sample
        msg = ""

    sigma_p2[Y] = Sigma

    cap_y = capA_ff[Y].reindex(elig).fillna(0.0).values if Y in capA_ff.columns else np.zeros(len(elig))
    total_cap = cap_y.sum()
    vw_weights_p2[Y] = cap_y / total_cap if total_cap > 0 else np.ones(len(elig)) / len(elig)
    inv_sets_p2[Y] = elig

    print(f"Y={Y}: eligible={len(elig)}{msg}")

vw_weight_dicts = {Y: {"isins": inv_sets_p2[Y], "weights": vw_weights_p2[Y]} for Y in inv_sets_p2}

print("\nSection 3.1 — Carbon intensity, WACI, Carbon Footprint")
rows_31 = {}
for Y in DECISION_YEARS:
    if Y not in inv_sets_p2 or Y not in part1_portfolios:
        continue

    # TRUE Part I MV portfolio
    isins_mv = list(part1_portfolios[Y]["isins"])
    w_mv = part1_portfolios[Y]["weights"]
    e_mv, r_mv, c_mv = get_carbon_vectors(Y, isins_mv)
    ci_mv = carbon_intensity_vec(e_mv, r_mv)

    # VW benchmark on the same eligible universe
    isins_vw = inv_sets_p2[Y]
    w_vw = vw_weights_p2[Y]
    e_vw, r_vw, c_vw = get_carbon_vectors(Y, isins_vw)
    ci_vw = carbon_intensity_vec(e_vw, r_vw)

    rows_31[Y] = {
        "WACI_MV": waci_metric(w_mv, ci_mv),
        "WACI_VW": waci_metric(w_vw, ci_vw),
        "CF_MV": cf_metric(w_mv, e_mv, c_mv),
        "CF_VW": cf_vw_metric(e_vw, c_vw),
    }

df_31 = pd.DataFrame(rows_31).T
df_31.index.name = "Year"
df_31.to_csv(f"{RESULTS_PART2}/carbon_metrics_mv_vw.csv")

waci_contrib_mv, waci_contrib_vw = {}, {}

for Y in DECISION_YEARS:
    if Y not in inv_sets_p2 or Y not in part1_portfolios:
        continue

    # MV contributors from the true Part I MV portfolio
    isins_mv = list(part1_portfolios[Y]["isins"])
    w_mv = part1_portfolios[Y]["weights"]
    e_mv, r_mv, _ = get_carbon_vectors(Y, isins_mv)
    ci_mv = carbon_intensity_vec(e_mv, r_mv)
    contrib_mv = pd.Series(np.nan_to_num(w_mv * ci_mv, nan=0.0), index=isins_mv)

    # VW contributors from benchmark weights
    isins_vw = inv_sets_p2[Y]
    w_vw = vw_weights_p2[Y]
    e_vw, r_vw, _ = get_carbon_vectors(Y, isins_vw)
    ci_vw = carbon_intensity_vec(e_vw, r_vw)
    contrib_vw = pd.Series(np.nan_to_num(w_vw * ci_vw, nan=0.0), index=isins_vw)

    for isin, val in contrib_mv.items():
        waci_contrib_mv.setdefault(isin, []).append(val)

    for isin, val in contrib_vw.items():
        waci_contrib_vw.setdefault(isin, []).append(val)

top10_mv = pd.Series({k: np.mean(v) for k, v in waci_contrib_mv.items()}).sort_values(ascending=False).head(10)
top10_vw = pd.Series({k: np.mean(v) for k, v in waci_contrib_vw.items()}).sort_values(ascending=False).head(10)
top10_mv_df = pd.DataFrame({"ISIN": top10_mv.index, "Name": top10_mv.index.map(firm_names), "Avg_WACI_Contribution_MV": top10_mv.values})
top10_vw_df = pd.DataFrame({"ISIN": top10_vw.index, "Name": top10_vw.index.map(firm_names), "Avg_WACI_Contribution_VW": top10_vw.values})
top10_mv_df.to_csv(f"{RESULTS_PART2}/top10_waci_contributors_mv.csv", index=False)
top10_vw_df.to_csv(f"{RESULTS_PART2}/top10_waci_contributors_vw.csv", index=False)

print("\nSection 3.2 — Min-variance with 50% Carbon Footprint constraint")
port_32, cf_32, waci_32 = {}, {}, {}
for Y in DECISION_YEARS:
    if Y not in inv_sets_p2:
        continue
    isins = inv_sets_p2[Y]
    Sigma = sigma_p2[Y]
    e, r, c = get_carbon_vectors(Y, isins)
    ec = e_over_c_vec(e, c)
    cf_mv_Y = df_31.loc[Y, "CF_MV"]
    if np.isnan(cf_mv_Y) or cf_mv_Y <= 0:
        print(f"Y={Y}: CF_MV invalid — skipping 3.2.")
        continue
    cf_target = 0.5 * cf_mv_Y
    w = solve_cvxpy(Sigma, ec, cf_target, mode="mv")
    if w is None:
        raise ValueError(f"Section 3.2 | Y={Y}: optimizer failed. Do not fallback silently.")
    achieved_cf = validate_cf_constraint(w, e, c, cf_target, "Section 3.2", Y)
    port_32[Y] = {"isins": isins, "weights": w}
    cf_32[Y] = achieved_cf
    waci_32[Y] = waci_metric(w, carbon_intensity_vec(e, r))

if port_32:
    pd.concat([
        pd.DataFrame({"Year": Y, "ISIN": v["isins"], "Weight": v["weights"]})
        for Y, v in port_32.items()
    ], ignore_index=True).to_csv(
        f"{RESULTS_PART2}/weights_32_mv_carbon05.csv",
        index=False
    )
else:
    print("Warning: port_32 is empty. No weights_32_mv_carbon05.csv file created.")

ret_32 = validate_oos_series(
    compute_oos_returns(port_32, ri_returns, parsed_dates, START_YEAR_OOS, END_YEAR_OOS),
    "P_mv_oos(0.5)"
)
ret_32.to_csv(f"{RESULTS_PART2}/returns_32_mv_carbon05.csv", header=["Return"])
stats_32 = pd.DataFrame({"P_mv_oos": perf_stats_extended(mv_r, rf=rf_monthly), "P_mv_oos(0.5)": perf_stats_extended(ret_32, rf=rf_monthly)}).T
stats_32.to_csv(f"{RESULTS_PART2}/stats_32_comparison.csv")

print("\nSection 3.3 — TE minimisation with 50% Carbon Footprint constraint")
port_33, cf_33, waci_33 = {}, {}, {}
for Y in DECISION_YEARS:
    if Y not in inv_sets_p2:
        continue
    isins = inv_sets_p2[Y]
    Sigma = sigma_p2[Y]
    e, r, c = get_carbon_vectors(Y, isins)
    ec = e_over_c_vec(e, c)
    w_vw = vw_weights_p2[Y]
    cf_vw_Y = df_31.loc[Y, "CF_VW"]
    if np.isnan(cf_vw_Y) or cf_vw_Y <= 0:
        print(f"Y={Y}: CF_VW invalid — skipping 3.3.")
        continue
    cf_target = 0.5 * cf_vw_Y
    w = solve_cvxpy(Sigma, ec, cf_target, w_ref=w_vw, mode="te")
    if w is None:
        raise ValueError(f"Section 3.3 | Y={Y}: optimizer failed. Do not fallback silently.")
    achieved_cf = validate_cf_constraint(w, e, c, cf_target, "Section 3.3", Y)
    port_33[Y] = {"isins": isins, "weights": w}
    cf_33[Y] = achieved_cf
    waci_33[Y] = waci_metric(w, carbon_intensity_vec(e, r))

if port_33:
    pd.concat([
        pd.DataFrame({"Year": Y, "ISIN": v["isins"], "Weight": v["weights"]})
        for Y, v in port_33.items()
    ], ignore_index=True).to_csv(
        f"{RESULTS_PART2}/weights_33_te_carbon05.csv",
        index=False
    )
else:
    print("Warning: port_33 is empty. No weights_33_te_carbon05.csv file created.")

ret_33 = validate_oos_series(
    compute_oos_returns(port_33, ri_returns, parsed_dates, START_YEAR_OOS, END_YEAR_OOS),
    "P_vw_oos(0.5)"
)
ret_33.to_csv(f"{RESULTS_PART2}/returns_33_te_carbon05.csv", header=["Return"])
stats_33 = pd.DataFrame({"P_vw_oos": perf_stats_extended(vw_r, rf=rf_monthly), "P_vw_oos(0.5)": perf_stats_extended(ret_33, rf=rf_monthly)}).T
stats_33.to_csv(f"{RESULTS_PART2}/stats_33_comparison.csv")

print(f"\nSection 4.1 — Net-zero portfolio (θ={THETA:.0%}/yr from Y0={Y0})")
isins_y0 = inv_sets_p2.get(Y0, [])
if isins_y0:
    e0, _, c0 = get_carbon_vectors(Y0, isins_y0)
    cf_vw_y0 = cf_vw_metric(e0, c0)
else:
    cf_vw_y0 = df_31.loc[Y0, "CF_VW"] if Y0 in df_31.index else np.nan
port_41, cf_41, waci_41 = {}, {}, {}
for Y in DECISION_YEARS:
    if Y not in inv_sets_p2:
        continue
    isins = inv_sets_p2[Y]
    Sigma = sigma_p2[Y]
    e, r, c = get_carbon_vectors(Y, isins)
    ec = e_over_c_vec(e, c)
    w_vw = vw_weights_p2[Y]
    cf_target = ((1.0 - THETA) ** (Y - Y0 + 1)) * cf_vw_y0
    w = solve_cvxpy(Sigma, ec, cf_target, w_ref=w_vw, mode="te")
    if w is None:
        raise ValueError(f"Section 4.1 | Y={Y}: optimizer failed. Do not fallback silently.")
    achieved_cf = validate_cf_constraint(w, e, c, cf_target, "Section 4.1", Y)
    port_41[Y] = {"isins": isins, "weights": w}
    cf_41[Y] = achieved_cf
    waci_41[Y] = waci_metric(w, carbon_intensity_vec(e, r))

if port_41:
    pd.concat([
        pd.DataFrame({"Year": Y, "ISIN": v["isins"], "Weight": v["weights"]})
        for Y, v in port_41.items()
    ], ignore_index=True).to_csv(
        f"{RESULTS_PART2}/weights_41_netzero.csv",
        index=False
    )
else:
    print("Warning: port_41 is empty. No weights_41_netzero.csv file created.")

ret_41 = validate_oos_series(
    compute_oos_returns(port_41, ri_returns, parsed_dates, START_YEAR_OOS, END_YEAR_OOS),
    "P_vw_oos(NZ)"
)
ret_41.to_csv(f"{RESULTS_PART2}/returns_41_netzero.csv", header=["Return"])
stats_41 = pd.DataFrame({
    "P_vw_oos": perf_stats_extended(vw_r, rf=rf_monthly),
    "P_vw_oos(0.5)": perf_stats_extended(ret_33, rf=rf_monthly),
    "P_vw_oos(NZ)": perf_stats_extended(ret_41, rf=rf_monthly),
}).T
stats_41.to_csv(f"{RESULTS_PART2}/stats_41_comparison.csv")

print("\nTotal attributed CO2 (V_2013 = $1 M starting wealth)…")
V0_M = 1.0
v_mv = annual_portfolio_values(mv_r, V0_M)
v_vw = annual_portfolio_values(vw_r, V0_M)
v_32 = annual_portfolio_values(ret_32, V0_M)
v_33 = annual_portfolio_values(ret_33, V0_M)
v_41 = annual_portfolio_values(ret_41, V0_M)
cf_mv_s = df_31["CF_MV"]
cf_vw_s = df_31["CF_VW"]
cf_32_s = pd.Series(cf_32).sort_index()
cf_33_s = pd.Series(cf_33).sort_index()
cf_41_s = pd.Series(cf_41).sort_index()

total_co2 = pd.DataFrame({
    "TotalCO2_MV": v_mv.reindex(df_31.index) * cf_mv_s,
    "TotalCO2_VW": v_vw.reindex(df_31.index) * cf_vw_s,
    "TotalCO2_MV05": v_32.reindex(pd.Index(sorted(cf_32_s.index))) * cf_32_s,
    "TotalCO2_TE05": v_33.reindex(pd.Index(sorted(cf_33_s.index))) * cf_33_s,
    "TotalCO2_NZ": v_41.reindex(pd.Index(sorted(cf_41_s.index))) * cf_41_s,
    "V_MV_MUSD": v_mv.reindex(df_31.index),
    "V_VW_MUSD": v_vw.reindex(df_31.index),
    "V_MV05_MUSD": v_32.reindex(pd.Index(sorted(cf_32_s.index))),
    "V_TE05_MUSD": v_33.reindex(pd.Index(sorted(cf_33_s.index))),
    "V_NZ_MUSD": v_41.reindex(pd.Index(sorted(cf_41_s.index))),
}).round(4)
total_co2.index.name = "Year"
total_co2.to_csv(f"{RESULTS_PART2}/total_attributed_co2.csv")

all_stats = pd.DataFrame({
    "P_mv_oos": perf_stats_extended(mv_r, rf=rf_monthly),
    "P_mv_oos(0.5)": perf_stats_extended(ret_32, rf=rf_monthly),
    "P_vw_oos": perf_stats_extended(vw_r, rf=rf_monthly),
    "P_vw_oos(0.5)": perf_stats_extended(ret_33, rf=rf_monthly),
    "P_vw_oos(NZ)": perf_stats_extended(ret_41, rf=rf_monthly),
}).T
all_stats.to_csv(f"{RESULTS_PART2}/all_portfolio_stats_extended.csv")

rel_stats = pd.DataFrame({
    "P_vw_oos(0.5)": perf_stats_relative(ret_33, vw_r, port_33, vw_weight_dicts, rf=rf_monthly),
    "P_vw_oos(NZ)": perf_stats_relative(ret_41, vw_r, port_41, vw_weight_dicts, rf=rf_monthly),
}).T
rel_stats.to_csv(f"{RESULTS_PART2}/relative_stats_te_portfolios.csv")

carbon_df = pd.DataFrame({
    "CF_MV": cf_mv_s,
    "CF_VW": cf_vw_s,
    "CF_MV_05": cf_32_s,
    "CF_TE_05": cf_33_s,
    "CF_NZ": cf_41_s,
    "WACI_MV": df_31["WACI_MV"],
    "WACI_VW": df_31["WACI_VW"],
    "WACI_MV_05": pd.Series(waci_32),
    "WACI_TE_05": pd.Series(waci_33),
    "WACI_NZ": pd.Series(waci_41),
})
if USE_LEDOIT_WOLF:
    carbon_df["LW_delta"] = pd.Series(lw_deltas)
    carbon_df["LW_r_bar"] = pd.Series(lw_r_bars)
carbon_df.rename_axis("Year").to_csv(f"{RESULTS_PART2}/all_carbon_metrics.csv")

nz_path = pd.Series({Y: ((1.0 - THETA) ** (Y - Y0 + 1)) * cf_vw_y0 for Y in DECISION_YEARS}, name="NZ_target").rename_axis("Year")
nz_path.to_csv(f"{RESULTS_PART2}/nz_target_path.csv")

pd.DataFrame({
    "P_mv_oos": perf_stats_extended(mv_r, rf=rf_monthly),
    "P_mv_oos(0.5)": perf_stats_extended(ret_32, rf=rf_monthly),
    "P_vw_oos": perf_stats_extended(vw_r, rf=rf_monthly),
    "P_vw_oos(0.5)": perf_stats_extended(ret_33, rf=rf_monthly),
}).T.to_csv(f"{RESULTS_PART2}/stats_34_section_comparison.csv")

pd.DataFrame({
    "P_vw_oos": perf_stats_extended(vw_r, rf=rf_monthly),
    "P_vw_oos(0.5)": perf_stats_extended(ret_33, rf=rf_monthly),
    "P_vw_oos(NZ)": perf_stats_extended(ret_41, rf=rf_monthly),
}).T.to_csv(f"{RESULTS_PART2}/stats_42_section_comparison.csv")

save_cumret_plot({"P_mv_oos": mv_r, "P_mv_oos(0.5)": ret_32}, f"{RESULTS_PART2}/cumret_32_mv_vs_mv05.png", "Section 3.2 — Cumulative Returns")
save_cumret_plot({"P_vw_oos": vw_r, "P_vw_oos(0.5)": ret_33}, f"{RESULTS_PART2}/cumret_33_vw_vs_te05.png", "Section 3.3 — Cumulative Returns")
save_cumret_plot({"P_vw_oos": vw_r, "P_vw_oos(0.5)": ret_33, "P_vw_oos(NZ)": ret_41}, f"{RESULTS_PART2}/cumret_41_42_vw_te05_nz.png", "Section 4.1 / 4.2 — Cumulative Returns")

cf_plot_df = pd.DataFrame({
    "CF_MV": cf_mv_s,
    "CF_MV_05": cf_32_s,
    "CF_VW": cf_vw_s,
    "CF_TE_05": cf_33_s,
    "CF_NZ": cf_41_s,
}).sort_index()
save_annual_line_plot(cf_plot_df, f"{RESULTS_PART2}/carbon_footprint_paths.png", "Carbon Footprint by Year", "Carbon Footprint (tCO2 / M USD invested)")

waci_plot_df = pd.DataFrame({
    "WACI_MV": df_31["WACI_MV"],
    "WACI_MV_05": pd.Series(waci_32),
    "WACI_VW": df_31["WACI_VW"],
    "WACI_TE_05": pd.Series(waci_33),
    "WACI_NZ": pd.Series(waci_41),
}).sort_index()
save_annual_line_plot(waci_plot_df, f"{RESULTS_PART2}/waci_paths.png", "WACI by Year", "WACI (tCO2 / M USD revenue)")

if port_32:
    export_weight_comparison(
        reference_portfolios=part1_portfolios,
        candidate_portfolios=port_32,
        output_csv=f"{RESULTS_PART2}/weights_comparison_32_vs_part1_mv.csv"
    )
    export_top_weight_changes(
        f"{RESULTS_PART2}/weights_comparison_32_vs_part1_mv.csv",
        f"{RESULTS_PART2}/top_weight_changes_32_vs_part1_mv.csv",
        top_n=20
    )
else:
    print("Warning: port_32 is empty. No comparison files created for Section 3.2.")

if port_33:
    export_weight_comparison(
        reference_portfolios=vw_weight_dicts,
        candidate_portfolios=port_33,
        output_csv=f"{RESULTS_PART2}/weights_comparison_33_vs_vw.csv"
    )
    export_top_weight_changes(
        f"{RESULTS_PART2}/weights_comparison_33_vs_vw.csv",
        f"{RESULTS_PART2}/top_weight_changes_33_vs_vw.csv",
        top_n=20
    )
else:
    print("Warning: port_33 is empty. No comparison files created for Section 3.3.")

if port_41:
    export_weight_comparison(
        reference_portfolios=vw_weight_dicts,
        candidate_portfolios=port_41,
        output_csv=f"{RESULTS_PART2}/weights_comparison_41_vs_vw.csv"
    )
    export_top_weight_changes(
        f"{RESULTS_PART2}/weights_comparison_41_vs_vw.csv",
        f"{RESULTS_PART2}/top_weight_changes_41_vs_vw.csv",
        top_n=20
    )
else:
    print("Warning: port_41 is empty. No comparison files created for Section 4.1.")

print("\nDone. Results written to:")
print("  ", RESULTS_PART1)
print("  ", RESULTS_PART2)


# Final implementation note:
# The code is written to preserve strict comparability between Part I and Part II
# while remaining consistent with the project brief. Any practical deviations from
# a textbook implementation — such as the custom covariance approximation or the
# handling of missing revenue in WACI — are disclosed explicitly in the report.