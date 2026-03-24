# ============================================================
# # ============================================================
# SAAM PROJECT 2026 — PART I
# Group assignment: Pacific region (PAC) | Scope 1 emissions
#
# Purpose
# -------
# This script implements Part I of the SAAM project by constructing:
#   1. A long-only minimum-variance portfolio
#   2. A value-weighted benchmark portfolio
#
# The portfolio formation follows the project brief:
# - monthly stock returns are computed from Datastream total return indices,
# - the investment set is defined at the end of each year Y,
# - moments are estimated from the previous 10 years of monthly returns,
# - the minimum-variance portfolio is optimized at the end of year Y
#   and held through year Y+1,
# - portfolio performance is computed monthly from January 2014 to
#   December 2025.
#
# Key implementation choices
# --------------------------
# 1. Region restriction:
#    Only firms from the assigned region (PAC) are eligible.
#
# 2. Shared investment universe for Parts I and II:
#    Firms without Scope 1 emissions data available at year-end Y are
#    excluded from the year Y investment set. This ensures consistency
#    between the standard portfolio analysis in Part I and the carbon-based
#    portfolio analysis in Part II.
#
# 3. Return estimation window:
#    Expected returns and the covariance matrix are estimated from the
#    previous 10 years (120 months) of monthly returns.
#
# 4. Covariance estimation:
#    The covariance matrix is the historical sample covariance matrix of
#    monthly returns over the estimation window, computed using pairwise
#    available observations. A very small diagonal ridge is added only to
#    improve numerical stability in the long-only optimization.
#
# 5. Missing-price treatment:
#    Datastream RI values below 0.5 are treated as missing. Internal gaps in
#    monthly RI levels are forward-filled only within the observed life of
#    the series. December missingness can be preserved for year-end
#    eligibility checks.
#
# 6. Delisting / terminal disappearance treatment:
#    If a firm economically disappears from the monthly RI series, a -100%
#    return is assigned at the terminal month and subsequent returns are set
#    to missing. Monthly market caps are also set to missing from that point
#    onward so that the benchmark does not retain stale positive weights.
#
# 7. Benchmark construction:
#    The value-weighted benchmark is computed within the same annual eligible
#    investment universe as the minimum-variance portfolio, using lagged
#    monthly market caps. This provides a like-for-like comparison and keeps
#    the benchmark consistent with the shared investment-set logic used in
#    the project.
#
# 8. Output:
#    The script exports:
#    - monthly out-of-sample returns,
#    - summary performance statistics,
#    - annual portfolio compositions,
#    - and a filled version of the official Part I Excel template.
# ============================================================

import os
import pandas as pd
import numpy as np
import re
from datetime import datetime
from scipy.optimize import minimize
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────
REGION_CODE         = "PAC"
LOW_PRICE_THRESHOLD = 0.5
RESULTS_DIR         = "resultsPart1"

try:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
except NameError:
    BASE_DIR = os.getcwd()

DATA_DIR = os.path.join(BASE_DIR, "data")


def data_path(filename: str) -> str:
    """
    Return the path to an input data file.

    Prefers the 'data' subfolder; falls back to the base directory so that
    both script-based and notebook-based execution can work.
    """
    candidate = os.path.join(DATA_DIR, filename)
    if os.path.exists(candidate):
        return candidate

    candidate = os.path.join(BASE_DIR, filename)
    if os.path.exists(candidate):
        return candidate

    raise FileNotFoundError(f"Could not find required file: {filename}")

WINDOW_YEARS    = 10
MIN_OBS_MONTHS  = 36
STALE_THRESHOLD = 0.50

START_YEAR_OOS = 2014
END_YEAR_OOS   = 2025

CO2_FILE = "DS_CO2_SCOPE_1_Y_2025.xlsx"
RF_FILE  = "Risk_Free_Rate_2025.xlsx"

SLSQP_MAXITER = 400
SLSQP_FTOL    = 1e-9

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS — loading & cleaning
# ─────────────────────────────────────────────────────────────────────────────

def load_datastream_wide(filepath, sheet=0):
    df = pd.read_excel(filepath, sheet_name=sheet, header=0, engine="openpyxl")
    df.columns = ["NAME", "ISIN"] + list(df.columns[2:])
    df = df[~df["NAME"].astype(str).str.startswith("$$ER", na=False)]
    df = df.dropna(subset=["ISIN"])
    df["ISIN"] = df["ISIN"].astype(str).str.strip()
    df["NAME"] = df["NAME"].astype(str).str.strip()
    df = df[df["ISIN"] != ""].set_index("ISIN")
    return df


def load_annual_panel(filepath: str) -> pd.DataFrame:
    """Read Datastream annual panel → ISIN-indexed DataFrame with int year cols."""
    df = pd.read_excel(filepath, engine="openpyxl")
    df.columns = ["NAME", "ISIN"] + list(df.columns[2:])
    df = df[~df["NAME"].astype(str).str.startswith("$$ER", na=False)]
    df = df.dropna(subset=["ISIN"])
    df["ISIN"] = df["ISIN"].astype(str).str.strip()
    df = df[df["ISIN"] != ""].set_index("ISIN").drop(columns=["NAME"], errors="ignore")
    df.columns = df.columns.astype(int)
    df = df.apply(pd.to_numeric, errors="coerce")
    return df


def load_rf_monthly(filepath):
    """
    Load Risk_Free_Rate_2025.xlsx.

    Format
    ------
    Index column : YYYYMM integer  (e.g. 202401 = January 2024)
    'RF' column  : monthly rate in PERCENT  (e.g. 0.41 = 0.41 %/month)

    Returns a pd.Series of monthly DECIMAL rates indexed by month-end
    Timestamps (same end-of-month convention as Datastream columns).
    Division by 100 converts percent → decimal.
    """
    df    = pd.read_excel(filepath, engine="openpyxl", index_col=0)
    df.index = df.index.astype(str).str.strip()
    dates = pd.to_datetime(df.index, format="%Y%m") + pd.offsets.MonthEnd(0)
    rf    = pd.to_numeric(df.iloc[:, 0], errors="coerce") / 100.0
    rf.index = dates
    return rf.dropna().sort_index()


def parse_monthly_columns(cols):
    parsed, keep = [], []
    for c in cols:
        if c == "NAME":
            continue
        dt = pd.to_datetime(str(c), errors="coerce", dayfirst=True)
        if pd.notna(dt):
            keep.append(c)
            parsed.append(pd.Timestamp(dt).normalize())
    return keep, parsed


def extract_delist_date(name_str):
    m = re.search(r"(?:DELIST|DEAD)\.(\d{2}/\d{2}/\d{2,4})", str(name_str))
    if not m:
        return None
    s = m.group(1)
    for fmt in ("%d/%m/%y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None


def forward_fill_middle_only(price_df, dates):
    arr = price_df.to_numpy(dtype=float, copy=True)
    for i in range(arr.shape[0]):
        row       = arr[i, :]
        valid_idx = np.where(~np.isnan(row))[0]
        if valid_idx.size == 0:
            continue
        first, last = valid_idx[0], valid_idx[-1]
        last_val = row[first]
        for k in range(first + 1, last + 1):
            if np.isnan(row[k]):
                row[k] = last_val
            else:
                last_val = row[k]
        arr[i, :] = row
    return pd.DataFrame(arr, index=price_df.index, columns=dates)


def clean_monthly_ri_prices(raw_ri_m, dates, low_price_threshold=0.5,
                             preserve_december_missing=True):
    """
    Clean monthly Datastream RI levels before return computation.

    Cleaning steps
    --------------
    1. Convert values to numeric.
    2. Treat RI values below the low-price threshold as missing.
    3. Forward-fill only internal gaps within the observed life of the series.
    4. Optionally preserve December missingness so that year-end eligibility
    checks are not distorted by temporary imputation.

    This function produces cleaned RI levels; returns are computed only after
    this cleaning step.
    """

    prices = raw_ri_m.copy().apply(pd.to_numeric, errors="coerce")
    prices = prices.mask(prices < low_price_threshold)
    if preserve_december_missing:
        dec_cols         = [d for d in dates if d.month == 12]
        orig_dec_missing = prices[dec_cols].isna()
    prices_filled = forward_fill_middle_only(prices, dates)
    if preserve_december_missing and len(dec_cols) > 0:
        prices_filled.loc[:, dec_cols] = (
            prices_filled.loc[:, dec_cols].mask(orig_dec_missing)
        )
    return prices_filled


def apply_delisting_to_returns(returns_df, price_df, delist_dates, dates):
    """
    Apply explicit Datastream delisting markers only when they are confirmed by
    disappearance in the monthly RI price panel.

    If the annual RI name contains a stale/inconsistent delisting marker but the
    monthly price series continues after the stated delisting date, the marker is
    ignored for return construction.
    """
    out = returns_df.copy()
    date_index = pd.Index(dates)

    for isin, ddate in delist_dates.items():
        if ddate is None or isin not in out.index or isin not in price_df.index:
            continue

        dts = pd.Timestamp(ddate).normalize()
        price_series = price_df.loc[isin]

        on_after_delist = date_index[date_index >= dts]
        if len(on_after_delist) == 0:
            continue
        default_dcol = on_after_delist.min()

        # Ignore stale annual markers if monthly prices continue after the stated delisting month
        future_window = date_index[date_index >= default_dcol]
        future_prices = price_series.loc[future_window]

        if future_prices.notna().any():
            if not (pd.isna(price_series.loc[default_dcol]) and future_prices.isna().all()):
                continue

        # Otherwise assign terminal loss at the first economically missing month
        prior_valid = date_index[(date_index < default_dcol) & price_series.loc[date_index].notna()]

        if len(prior_valid) > 0:
            last_valid = prior_valid.max()
            candidate_window = date_index[(date_index > last_valid) & (date_index <= default_dcol)]
            missing = candidate_window[price_series.loc[candidate_window].isna()]
            dcol = missing.min() if len(missing) > 0 else default_dcol
        else:
            dcol = default_dcol

        out.at[isin, dcol] = -1.0
        after = date_index[date_index > dcol]
        if len(after) > 0:
            out.loc[isin, after] = np.nan

    return out
    #fix the bug with delisted
def apply_terminal_loss_for_permanent_disappearances(returns_df, price_df, dates):
    """
    Assign a terminal -100% return to firms that permanently disappear from
    the cleaned monthly RI series.

    If a firm's cleaned RI level becomes missing and remains missing until the
    end of the sample, the first permanently missing month is treated as a
    terminal loss month. The return at that month is set to -100% if it is
    still missing, and all subsequent returns are set to missing.
    """
    out = returns_df.copy()
    date_index = pd.Index(dates)

    for isin in out.index:
        if isin not in price_df.index:
            continue

        price_series = price_df.loc[isin]
        valid_dates = date_index[price_series.loc[date_index].notna()]

        if len(valid_dates) == 0:
            continue

        last_valid_price_date = valid_dates.max()
        tail_dates = date_index[date_index > last_valid_price_date]

        if len(tail_dates) == 0:
            continue

        if price_series.loc[tail_dates].isna().all():
            first_missing = tail_dates.min()

            if pd.isna(out.at[isin, first_missing]):
                out.at[isin, first_missing] = -1.0

            later = date_index[date_index > first_missing]
            if len(later) > 0:
                out.loc[isin, later] = np.nan

    return out
    #current missing fix
def adjust_mv_caps_for_terminal_events(mv_caps, ri_returns, dates):
    """
    Remove stale benchmark market-cap weights after terminal-loss events.

    If a firm experiences a terminal-loss event in RI returns (a -100% return
    followed only by missing returns), its monthly market capitalization is
    set to missing from that terminal month onward. This prevents the
    value-weighted benchmark from retaining stale positive weights in
    economically dead securities.
    """
    caps = mv_caps.copy()
    date_index = pd.Index(dates)

    for isin in caps.index.intersection(ri_returns.index):
        series = ri_returns.loc[isin, date_index]
        minus100_dates = date_index[series == -1.0]

        if len(minus100_dates) == 0:
            continue

        for d in minus100_dates:
            after = date_index[date_index > d]

            if len(after) == 0 or series.loc[after].isna().all():
                caps.loc[isin, date_index[date_index >= d]] = np.nan
                break

    return caps    
# ─────────────────────────────────────────────────────────────────────────────
# INVESTMENT SET, MOMENTS, OPTIMISATION
# ─────────────────────────────────────────────────────────────────────────────

def year_end_col(dates, year):
    decs = [d for d in dates if d.year == year and d.month == 12]
    return max(decs) if decs else None


def window_cols(dates, year_end, window_years=10):
    start = pd.Timestamp(f"{year_end - window_years + 1}-01-01")
    end   = pd.Timestamp(f"{year_end}-12-31")
    return [d for d in dates if start <= d <= end]


def stale_mask(returns_df, cols, threshold=0.50):
    sub   = returns_df[cols]
    denom = sub.notna().sum(axis=1).replace(0, np.nan)
    frac0 = (sub == 0).sum(axis=1) / denom
    return frac0 > threshold


def build_investment_set(ri_prices, ri_returns, dates, year_end,
                         co2_panel=None):
    """
    Construct the eligible investment set at the end of year Y.

    A firm is eligible if all of the following conditions are satisfied:
      1. Its December RI value is observed and passes the low-price screen.
      2. It has at least MIN_OBS_MONTHS valid monthly returns in the
         previous 10-year estimation window.
      3. It is not classified as a stale-price stock based on the proportion
         of zero monthly returns in the same estimation window.
      4. If a carbon panel is supplied, the firm has Scope 1 data available
         at year-end Y.

    The carbon screen is applied already in Part I so that the investment
    universe remains consistent with Part II of the project. Annual carbon
    data are forward-filled before this function is called, in line with the
    project rule for missing values in the middle or at the end of the sample.
    """
    dec_col = year_end_col(dates, year_end)
    if dec_col is None:
        raise ValueError(f"No December column for year {year_end}.")

    cols     = window_cols(dates, year_end, WINDOW_YEARS)
    ok_price = ri_prices[dec_col].notna()
    n_obs    = ri_returns[cols].notna().sum(axis=1)
    ok_obs   = n_obs >= MIN_OBS_MONTHS
    ok_stale = ~stale_mask(ri_returns, cols, STALE_THRESHOLD).fillna(True)

    ok_carbon = pd.Series(True, index=ri_prices.index)
    if co2_panel is not None:
        if year_end in co2_panel.columns:
            ok_carbon = (
                co2_panel[year_end]
                .reindex(ri_prices.index)
                .notna()
                .fillna(False)
            )
        else:
            print(f"  Warning: year {year_end} not in CO2 panel — "
                  "carbon filter skipped.")

    eligible = ri_prices.index[ok_price & ok_obs & ok_stale & ok_carbon]
    return list(eligible), cols


def estimate_moments(ri_returns, isins, cols, ridge=1e-8):
    """
    Estimate expected returns and the covariance matrix on the 10-year
    monthly estimation window.

    Mean estimator
    --------------
    Expected returns are estimated as the arithmetic mean of observed
    monthly returns over the window.

    Covariance estimator
    --------------------
    The covariance matrix is estimated as the historical sample covariance
    matrix of monthly returns over the same window, using pairwise available
    observations.

    Numerical stability
    -------------------
    A very small diagonal ridge is added to the covariance matrix solely to
    stabilize the numerical optimization. This ridge is not intended to alter
    the economic interpretation of the estimator.
    """

    R = ri_returns.loc[isins, cols].astype(float)

    # Mean monthly returns
    mu = R.mean(axis=1, skipna=True).to_numpy()

    # Sample covariance across assets
    # R has shape (N assets, T months), so transpose to (T months, N assets)
    R_np = R.to_numpy()  # shape (N, T)
    mu_col = np.nanmean(R_np, axis=1, keepdims=True)
    demeaned = np.where(np.isnan(R_np), 0.0, R_np - mu_col)
    valid_counts = np.isfinite(R_np).sum(axis=1)
    tau = valid_counts.min()  # or use fixed 120
    Sigma = demeaned @ demeaned.T / tau
    
    # Safety check
    if np.isnan(Sigma).any():
        raise ValueError("Covariance matrix contains NaN values after estimation.")

    # Tiny ridge for numerical stability
    Sigma = Sigma + ridge * np.eye(Sigma.shape[0])

    return mu, Sigma


def solve_min_variance(Sigma):
    """Long-only min-var: min a'Σa  s.t. sum(a)=1, a≥0."""
    n  = Sigma.shape[0]
    a0 = np.full(n, 1.0 / n)
    res = minimize(
        lambda a: float(a @ Sigma @ a), a0,
        method="SLSQP",
        bounds=[(0.0, 1.0)] * n,
        constraints=[{"type": "eq", "fun": lambda a: np.sum(a) - 1.0}],
        options={"maxiter": SLSQP_MAXITER, "ftol": SLSQP_FTOL, "disp": False}
    )
    if not res.success:
        raise RuntimeError(f"SLSQP failed: {res.message}")
    w = np.where(res.x < 1e-10, 0.0, res.x)
    return w / w.sum()


def compute_mv_oos_returns(portfolios, ri_returns, dates):
    """
    Compute monthly out-of-sample returns for the minimum-variance portfolio.

    Important implementation rule:
    - A missing realized return is NOT silently treated as 0 if the stock still
      has a non-trivial portfolio weight.
    - Missing returns are allowed only for positions whose weight is already
      economically zero (for example after a captured -100% delisting event).
    """
    out_r, out_d = [], []
    WEIGHT_TOL = 1e-12

    for year_end, info in portfolios.items():
        invest_year = year_end + 1
        year_months = [d for d in dates
                       if pd.Timestamp(f"{invest_year}-01-01") <= d
                       <= pd.Timestamp(f"{invest_year}-12-31")]
        if not year_months:
            continue

        isins = info["isins"]
        w = info["weights"].copy()

        for d in year_months:
            r_series = ri_returns.loc[isins, d].copy()

            # Missing realized returns are only acceptable if the current
            # portfolio weight is already economically zero.
            missing_mask = r_series.isna().to_numpy()
            bad_mask = missing_mask & (w > WEIGHT_TOL)

            if bad_mask.any():
                bad_isins = list(np.array(isins)[bad_mask])
                bad_weights = w[bad_mask]
                raise ValueError(
                    f"Missing realized returns for invested names on {d.date()}. "
                    f"ISINs={bad_isins}, weights={bad_weights.tolist()}"
                )

            # Safe only for names whose weight is already zero / negligible
            r_i = r_series.fillna(0.0).to_numpy()

            r_p = float(w @ r_i)
            out_r.append(r_p)
            out_d.append(d)

            denom = 1.0 + r_p
            if denom <= 0:
                raise ValueError(
                    f"Portfolio value collapsed to non-positive level on {d.date()} "
                    f"(1 + r_p = {denom})."
                )

            w = w * (1.0 + r_i) / denom
            w = np.where(w < WEIGHT_TOL, 0.0, w)

            if w.sum() <= 0:
                raise ValueError(f"Portfolio weights sum to zero after update on {d.date()}.")

            w = w / w.sum()

    return pd.Series(out_r, index=pd.to_datetime(out_d)).sort_index()

# Methodology note:
# The value-weighted benchmark is constructed within the same annual eligible
# investment universe as the minimum-variance portfolio. At the end of each
# year Y, the eligible set is defined using the project screens: region
# membership, sufficient return history, exclusion of stale-price firms, and
# availability of Scope 1 emissions data at year-end Y.
#
# For the subsequent year Y+1, benchmark returns are computed monthly using
# lagged month-end market capitalizations within that eligible set only.
# This choice provides a like-for-like comparison with the optimized portfolio
# and keeps the benchmark consistent with the shared investment-set logic used
# later in Part II.
def compute_vw_oos_returns(portfolios, ri_returns, mv_caps, dates,
                           start_year=2014, end_year=2025):
    """

    Compute the out-of-sample value-weighted benchmark returns.

    Methodology
    -----------
    For calendar year Y+1, the benchmark universe is restricted to the same
    eligible investment set used for the minimum-variance portfolio at the
    end of year Y. Within that eligible set, monthly benchmark weights are
    based on lagged month-end market capitalizations.

    Missing-data treatment
    ----------------------
    - cap present & return present   -> include
    - cap missing & return missing   -> exclude
    - cap present & return missing   -> error
    - cap missing & return present   -> error

    This implementation is intended to preserve economic consistency and to
    provide a like-for-like comparison with the minimum-variance portfolio.
    """

    out_r, out_d = [], []
    dates_sorted = sorted(dates)
    pos = {d: i for i, d in enumerate(dates_sorted)}

    for year_end, info in portfolios.items():
        invest_year = year_end + 1
        if invest_year < start_year or invest_year > end_year:
            continue

        eligible_isins = list(info["isins"])

        year_months = [d for d in dates_sorted
                       if pd.Timestamp(f"{invest_year}-01-01") <= d
                       <= pd.Timestamp(f"{invest_year}-12-31")]

        for d in year_months:
            i = pos.get(d)
            if i is None or i == 0:
                continue

            d_prev = dates_sorted[i - 1]

            if d_prev not in mv_caps.columns:
                raise ValueError(f"Lagged market-cap column missing for benchmark month {d_prev.date()}")
            if d not in ri_returns.columns:
                raise ValueError(f"Return column missing for benchmark month {d.date()}")

            caps = mv_caps.reindex(eligible_isins)[d_prev]
            rets = ri_returns.reindex(eligible_isins)[d]

            caps_missing = caps.isna()
            rets_missing = rets.isna()

            # True inconsistency = one side missing, the other present
            inconsistent = caps_missing ^ rets_missing
            if inconsistent.any():
                bad_isins = list(caps.index[inconsistent])
                raise ValueError(
                    f"VW benchmark inconsistent data on {d.date()}.\n"
                    f"Dropped ISINs: {bad_isins}\n"
                    f"Reason: market cap and return availability do not match."
                )

            # Valid names for the benchmark this month
            valid = (~caps_missing) & (~rets_missing)

            if valid.sum() == 0:
                raise ValueError(f"No valid benchmark constituents on {d.date()}")

            cap_sum = caps.loc[valid].sum()
            if pd.isna(cap_sum) or cap_sum <= 0:
                raise ValueError(f"Invalid lagged market-cap sum for benchmark on {d.date()}")

            w = caps.loc[valid] / cap_sum
            r_p = float(w.to_numpy() @ rets.loc[valid].to_numpy())

            out_r.append(r_p)
            out_d.append(d)

    return pd.Series(out_r, index=pd.to_datetime(out_d)).sort_index()


def perf_stats(r, rf=None):
    """
    Template-aligned annualized performance statistics.
    """
    r = r.dropna()
    T = len(r)

    if T == 0:
        return {
            "Annualized Average Return": np.nan,
            "Annualized Volatility": np.nan,
            "Annualized Cumulative Return": np.nan,
            "Sharpe Ratio": np.nan,
            "Minimum Monthly Return": np.nan,
            "Maximum Monthly Return": np.nan,
        }

    rf_aligned = rf.reindex(r.index).fillna(0.0) if rf is not None else pd.Series(0.0, index=r.index)

    excess = r - rf_aligned
    mean_m = r.mean()
    vol_m = r.std(ddof=1)
    vol_excess = excess.std(ddof=1)

    return {
        "Annualized Average Return": 12.0 * mean_m,
        "Annualized Volatility": vol_m * np.sqrt(12.0),
        "Annualized Cumulative Return": (1.0 + r).prod() ** (12.0 / T) - 1.0,
        "Sharpe Ratio": (excess.mean() / vol_excess) * np.sqrt(12.0) if vol_excess > 0 else np.nan,
        "Minimum Monthly Return": r.min(),
        "Maximum Monthly Return": r.max(),
    }
# summary stats fill
def export_part1_excel_template(template_path, output_path, stats, out_df):
    """
    Fill the official Part I SAAM Excel template.

    Expected structure in template:
    - B3:B8   = Value-weighted stats
    - C3:C8   = Minimum-variance stats
    - F3:F146 = Value-weighted monthly returns
    - G3:G146 = Minimum-variance monthly returns
    - Plot inserted around B10
    """

    wb = load_workbook(template_path)
    ws = wb["Sheet1"]

    # ---------------------------------------------------------
    # 1) Write summary statistics
    # ---------------------------------------------------------
    # Column B = Value-weighted portfolio
    ws["B3"] = float(stats.loc["Value Weighted", "Annualized Average Return"])
    ws["B4"] = float(stats.loc["Value Weighted", "Annualized Volatility"])
    ws["B5"] = float(stats.loc["Value Weighted", "Annualized Cumulative Return"])
    ws["B6"] = float(stats.loc["Value Weighted", "Sharpe Ratio"])
    ws["B7"] = float(stats.loc["Value Weighted", "Minimum Monthly Return"])
    ws["B8"] = float(stats.loc["Value Weighted", "Maximum Monthly Return"])

    # Column C = Minimum-variance portfolio
    ws["C3"] = float(stats.loc["Minimum Variance", "Annualized Average Return"])
    ws["C4"] = float(stats.loc["Minimum Variance", "Annualized Volatility"])
    ws["C5"] = float(stats.loc["Minimum Variance", "Annualized Cumulative Return"])
    ws["C6"] = float(stats.loc["Minimum Variance", "Sharpe Ratio"])
    ws["C7"] = float(stats.loc["Minimum Variance", "Minimum Monthly Return"])
    ws["C8"] = float(stats.loc["Minimum Variance", "Maximum Monthly Return"])

    # ---------------------------------------------------------
    # 2) Write monthly returns
    # ---------------------------------------------------------
    if len(out_df) != 144:
        raise ValueError(f"Expected 144 monthly observations, got {len(out_df)}")

    start_row = 3
    for i, row in enumerate(out_df.itertuples(index=False), start=start_row):
        ws[f"F{i}"] = float(row.VW_Return)
        ws[f"G{i}"] = float(row.MV_Return)

    # ---------------------------------------------------------
    # 3) Create cumulative return plot
    # ---------------------------------------------------------
    plot_path = os.path.join(os.path.dirname(output_path), "part1_cumulative_plot.png")

    plt.figure(figsize=(9, 4.8))
    plt.plot(out_df["Date"], out_df["VW_CumReturn"], label="Value-weighted portfolio")
    plt.plot(out_df["Date"], out_df["MV_CumReturn"], label="Minimum-variance portfolio")
    plt.title("Cumulative Returns (2014–2025)")
    plt.xlabel("Date")
    plt.ylabel("Growth of $1")
    plt.legend()
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(plot_path, dpi=200, bbox_inches="tight")
    plt.close()

    # ---------------------------------------------------------
    # 4) Insert plot into template
    # ---------------------------------------------------------
    img = XLImage(plot_path)
    img.width = 520
    img.height = 280
    ws.add_image(img, "B10")

    # ---------------------------------------------------------
    # 5) Save filled workbook
    # ---------------------------------------------------------
    wb.save(output_path)
# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("SAAM 2026 — PART I (v6)")
    print(f"Region: {REGION_CODE}")
    print("=" * 60)
    os.makedirs(RESULTS_DIR, exist_ok=True)

    # ── Static ───────────────────────────────────────────────────────────────
    static         = pd.read_excel(data_path("Static_2025.xlsx"), engine="openpyxl")
    static.columns = ["ISIN", "NAME", "Country", "Region"]
    static["ISIN"] = static["ISIN"].astype(str).str.strip()
    pac            = static[static["Region"] == REGION_CODE].copy().set_index("ISIN")
    pac_isins      = set(pac.index)
    print(f"Pacific firms in Static: {len(pac_isins)}")

    # ── Annual RI (delist dates) ──────────────────────────────────────────────
    ri_y         = load_datastream_wide(data_path("DS_RI_T_USD_Y_2025.xlsx"))
    ri_y         = ri_y[ri_y.index.isin(pac_isins)].copy()
    delist_dates = {isin: extract_delist_date(ri_y.at[isin, "NAME"])
                    for isin in ri_y.index}
    print(f"Delisted Pacific firms: {sum(v is not None for v in delist_dates.values())}")

    # ── Monthly RI ────────────────────────────────────────────────────────────
    ri_m_raw                = load_datastream_wide(data_path("DS_RI_T_USD_M_2025.xlsx"))
    keep_cols, parsed_dates = parse_monthly_columns(list(ri_m_raw.columns))
    ri_m                    = ri_m_raw[["NAME"] + keep_cols].copy()
    ri_m.columns            = ["NAME"] + parsed_dates
    ri_m                    = ri_m[ri_m.index.isin(pac_isins)].copy()
    parsed_dates            = [d for d in parsed_dates
                                if d <= pd.Timestamp("2025-12-31")]
    ri_m                    = ri_m[["NAME"] + parsed_dates]
    print(f"Monthly RI: {ri_m.shape[0]} firms, {len(parsed_dates)} months")

    # ── Monthly MV ────────────────────────────────────────────────────────────
    mv_m_raw           = load_datastream_wide(data_path("DS_MV_T_USD_M_2025.xlsx"))
    keep_cols_mv, _mv  = parse_monthly_columns(list(mv_m_raw.columns))
    mv_m               = mv_m_raw[["NAME"] + keep_cols_mv].copy()
    mv_m.columns       = ["NAME"] + _mv
    mv_m               = mv_m[mv_m.index.isin(pac_isins)].copy()
    mv_m               = mv_m[["NAME"] + parsed_dates].copy()
    mv_m[parsed_dates] = mv_m[parsed_dates].apply(pd.to_numeric, errors="coerce")
    print(f"Monthly MV: {mv_m.shape[0]} firms")

    # ── Common ISINs ──────────────────────────────────────────────────────────
    common_isins = (pac_isins
                    .intersection(set(ri_m.index))
                    .intersection(set(mv_m.index)))
    ri_m = ri_m.loc[list(common_isins)].copy()
    mv_m = mv_m.loc[list(common_isins)].copy()
    print(f"Common ISINs: {len(common_isins)}")

        # ── Clean prices & returns ────────────────────────────────────────────────
    # Two RI price panels:
    # 1) ri_prices_universe  -> preserves December missingness for year-end eligibility
    # 2) ri_prices_returns   -> does NOT preserve December missingness, so temporary
    #    December quote gaps do not create artificial missing realized returns
    ri_prices_universe = clean_monthly_ri_prices(
        ri_m[parsed_dates], parsed_dates, LOW_PRICE_THRESHOLD, preserve_december_missing=True
    )
    ri_prices_returns = clean_monthly_ri_prices(
        ri_m[parsed_dates], parsed_dates, LOW_PRICE_THRESHOLD, preserve_december_missing=False
    )

    all_missing = ri_prices_returns.isna().all(axis=1)
    if all_missing.any():
        missing_isins = ri_prices_returns.index[all_missing]
        ri_prices_returns = ri_prices_returns.loc[~all_missing].copy()
        ri_prices_universe = ri_prices_universe.loc[~all_missing].copy()
        mv_m = mv_m.drop(index=missing_isins, errors="ignore")
        print(f"Dropped fully-missing RI rows: {all_missing.sum()}")

    ri_returns = ri_prices_returns.pct_change(axis=1, fill_method=None)

    ri_returns = apply_delisting_to_returns(
        ri_returns,
        ri_prices_returns,
        {k: v for k, v in delist_dates.items() if k in ri_returns.index},
        parsed_dates
    )

    ri_returns = apply_terminal_loss_for_permanent_disappearances(
        ri_returns,
        ri_prices_returns,
        parsed_dates
    )

    mv_caps_adjusted = adjust_mv_caps_for_terminal_events(
        mv_m[parsed_dates],
        ri_returns,
        parsed_dates
    )
     

    # Forward-fill annual CO2 data before building the investment set.
    # This implements the project rule that missing annual carbon values may be
    # carried forward when the gap occurs in the middle of the sample or at the
    # end of the sample. It also ensures that Parts I and II use the same annual
    # eligibility logic.
    print(f"Loading CO2 panel: {CO2_FILE}")
    co2_panel = load_annual_panel(data_path(CO2_FILE))
    co2_panel = co2_panel.T.ffill().T          # ← Bug 1 fix
    print(f"  {co2_panel.shape[0]} ISINs, "
          f"years {co2_panel.columns.min()}–{co2_panel.columns.max()}")

    # ── Load risk-free rate ───────────────────────────────────────────────────
    print(f"Loading risk-free rate: {RF_FILE}")
    try:
        rf_monthly = load_rf_monthly(data_path(RF_FILE))
        print(f"  {len(rf_monthly)} obs, "
              f"mean={rf_monthly.mean()*100:.3f} %/month")
    except Exception as exc:
        print(f"  WARNING: could not load RF ({exc}). Using rf=0.")
        rf_monthly = None

    # ── Build annual MV portfolios ────────────────────────────────────────────
    portfolios = {}
    for year_end in range(START_YEAR_OOS - 1, END_YEAR_OOS):   # 2013 … 2024
        elig, cols = build_investment_set(
            ri_prices_universe, ri_returns, parsed_dates, year_end,
            co2_panel=co2_panel
        )
        if len(elig) < 2:
            print(f"Year {year_end}: eligible={len(elig)} (skipped)")
            continue
        _, Sigma = estimate_moments(ri_returns, elig, cols)    # 10-year sample covariance estimate
        w        = solve_min_variance(Sigma)
        portfolios[year_end] = {"isins": elig, "weights": w}
        print(f"Year {year_end}: eligible={len(elig)}, max_w={w.max():.4f}")

    # ── OOS returns 2014–2025 ─────────────────────────────────────────────────
    mv_r = compute_mv_oos_returns(portfolios, ri_returns, parsed_dates)
    vw_r = compute_vw_oos_returns(
        portfolios, ri_returns, mv_caps_adjusted, parsed_dates,
        START_YEAR_OOS, END_YEAR_OOS
    )

    # Convert Datastream monthly dates (often last trading day) to true calendar month-end
    mv_r.index = pd.to_datetime(mv_r.index) + pd.offsets.MonthEnd(0)
    vw_r.index = pd.to_datetime(vw_r.index) + pd.offsets.MonthEnd(0)

    # In case multiple original dates map to the same calendar month-end, keep the last
    mv_r = mv_r[~mv_r.index.duplicated(keep="last")]
    vw_r = vw_r[~vw_r.index.duplicated(keep="last")]

    # Force exact required OOS monthly index: Jan-2014 to Dec-2025 (144 months)
    expected_months = pd.date_range(
        start=f"{START_YEAR_OOS}-01-31",
        end=f"{END_YEAR_OOS}-12-31",
        freq="M"
    )

    mv_r = mv_r.reindex(expected_months)
    vw_r = vw_r.reindex(expected_months)

    # Hard fail if any required month is missing
    if mv_r.isna().any():
        missing_mv = mv_r.index[mv_r.isna()]
        raise ValueError(
            "Minimum-variance portfolio is missing OOS returns for these months: "
            f"{[d.strftime('%Y-%m-%d') for d in missing_mv]}"
        )

    if vw_r.isna().any():
        missing_vw = vw_r.index[vw_r.isna()]
        raise ValueError(
            "Value-weighted benchmark is missing OOS returns for these months: "
            f"{[d.strftime('%Y-%m-%d') for d in missing_vw]}"
        )

    if len(mv_r) != 144 or len(vw_r) != 144:
        raise ValueError(
            f"Unexpected OOS length: MV={len(mv_r)}, VW={len(vw_r)}. Expected 144 months."
        )

    # ── Export ────────────────────────────────────────────────────────────────
    out = pd.DataFrame({
        "Date": expected_months,
        "MV_Return": mv_r.values,
        "VW_Return": vw_r.values,
    })
    out["MV_CumReturn"] = (1 + out["MV_Return"]).cumprod()
    out["VW_CumReturn"] = (1 + out["VW_Return"]).cumprod()
    out.to_csv(os.path.join(RESULTS_DIR, "part1_results.csv"), index=False)

    stats = pd.DataFrame({
        "Minimum Variance": perf_stats(mv_r, rf=rf_monthly),
        "Value Weighted": perf_stats(vw_r, rf=rf_monthly),
    }).T
    stats.to_csv(os.path.join(RESULTS_DIR, "part1_summary_statistics.csv"))

        # ── Fill official Part I Excel template ───────────────────────────────
    template_input = data_path("Template for Part I-SAAM.xlsx")
    template_output = os.path.join(RESULTS_DIR, "Template_for_Part_I_SAAM_FILLED.xlsx")

    export_part1_excel_template(
        template_path=template_input,
        output_path=template_output,
        stats=stats,
        out_df=out
    )

    comp = []
    for year_end, info in portfolios.items():
        comp.append(pd.DataFrame({
            "Year": year_end + 1,
            "ISIN": info["isins"],
            "Weight": info["weights"],
        }))
    pd.concat(comp, ignore_index=True).to_csv(
        os.path.join(RESULTS_DIR, "part1_portfolio_compositions.csv"),
        index=False
    )

    print("=" * 60)
    print("Done. Results written to:", RESULTS_DIR)
    print("  part1_results.csv")
    print("  part1_summary_statistics.csv")
    print("  part1_portfolio_compositions.csv")
    print("  Template_for_Part_I_SAAM_FILLED.xlsx")
    print("=" * 60)


if __name__ == "__main__":
    main()
