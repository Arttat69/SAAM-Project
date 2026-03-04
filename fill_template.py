# ============================================================
# TEMPLATE FILLER - PART I SAAM PROJECT
# Automatically fills Template-for-Part-I-SAAM.xlsx with results
# ============================================================

import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.cell.cell import MergedCell
from datetime import datetime

print("=" * 60)
print("TEMPLATE FILLER FOR PART I")
print("=" * 60)

# directory where Part I results CSVs are stored
RESULTS_DIR = "resultsPart1"

# ============================================================
# 1. LOAD RESULTS
# ============================================================

print("\n[1] Loading results from CSV files...")

results_path = os.path.join(RESULTS_DIR, "part1_results.csv")
summary_path = os.path.join(RESULTS_DIR, "part1_summary_statistics.csv")

try:
    results = pd.read_csv(results_path)
    results['Date'] = pd.to_datetime(results['Date'])
    print(f"   ✓ Loaded results from '{results_path}': {len(results)} months")
except FileNotFoundError:
    print(f"   ✗ Error: '{results_path}' not found!")
    print("   Make sure you have run the Part I script and that the resultsPart1 folder exists.")
    exit(1)

try:
    summary = pd.read_csv(summary_path, index_col=0)
    print(f"   ✓ Loaded summary statistics from '{summary_path}'")
except FileNotFoundError:
    print(f"   ✗ Error: '{summary_path}' not found!")
    exit(1)

# ============================================================
# 2. LOAD TEMPLATE
# ============================================================

print("\n[2] Loading template...")

try:
    wb = openpyxl.load_workbook('Template for Part I-SAAM.xlsx')
    print(f"   ✓ Template loaded")
    print(f"   Sheets found: {wb.sheetnames}")
except FileNotFoundError:
    print("   ✗ Error: Template-for-Part-I-SAAM.xlsx not found!")
    exit(1)

# Get the active sheet (assuming there's one main sheet)
ws = wb.active
sheet_name = ws.title
print(f"   Working with sheet: '{sheet_name}'")

# ============================================================
# 3. ANALYZE TEMPLATE STRUCTURE
# ============================================================

print("\n[3] Analyzing template structure...")

# Read first 20 rows to understand structure
template_structure = []
for row_idx in range(1, 21):
    row_values = []
    for col_idx in range(1, 10):  # Check first 9 columns
        cell = ws.cell(row_idx, col_idx)
        if cell.value is not None:
            row_values.append((col_idx, cell.value))
    if row_values:
        template_structure.append((row_idx, row_values))

print("\n   Template preview (first 20 rows):")
for row_idx, values in template_structure[:20]:
    print(f"   Row {row_idx}: {values}")

# ============================================================
# 4. DETERMINE FILLING STRATEGY
# ============================================================

print("\n[4] Determining where to place data...")

# Common template patterns:
# Pattern A: Headers in row 1, data starts row 2
# Pattern B: Title/info at top, headers around row 5-10
# Pattern C: Summary stats section, then monthly data section

# Strategy: Find first empty area or data section
# Look for keywords like "Date", "Return", "Month", "Portfolio"

header_row = None
data_start_row = None

for row_idx in range(1, 50):
    row_text = ' '.join([str(ws.cell(row_idx, col).value or '') for col in range(1, 10)])
    row_text_lower = row_text.lower()
    
    if 'date' in row_text_lower or 'month' in row_text_lower:
        header_row = row_idx
        data_start_row = row_idx + 1
        print(f"   Found data section at row {header_row} (header)")
        break

if header_row is None:
    # Default: assume standard layout with header in row 1
    header_row = 1
    data_start_row = 2
    print(f"   Using default layout (header row 1, data row 2)")

# ============================================================
# 5. FILL MONTHLY RETURNS
# ============================================================

print("\n[5] Filling monthly returns...")

# Column mapping (adjust based on template)
col_date = 1  # Column A
col_mv_return = 2  # Column B
col_vw_return = 3  # Column C

# Ensure data_start_row is not inside a merged region
while True:
    coords = [
        ws.cell(data_start_row, col_date).coordinate,
        ws.cell(data_start_row, col_mv_return).coordinate,
        ws.cell(data_start_row, col_vw_return).coordinate,
    ]
    in_merged = any(
        coord in merged_range
        for merged_range in ws.merged_cells.ranges
        for coord in coords
    )
    if in_merged:
        data_start_row += 1
    else:
        break

# Write headers (if not already present)
ws.cell(header_row, col_date, "Date")
ws.cell(header_row, col_mv_return, "MV Portfolio Return")
ws.cell(header_row, col_vw_return, "VW Portfolio Return")

# Format headers
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

for col in [col_date, col_mv_return, col_vw_return]:
    cell = ws.cell(header_row, col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Write data
for idx, row in results.iterrows():
    excel_row = data_start_row + idx

    # Skip down if this row's target cells are part of a merged region
    while (
        isinstance(ws.cell(excel_row, col_date), MergedCell)
        or isinstance(ws.cell(excel_row, col_mv_return), MergedCell)
        or isinstance(ws.cell(excel_row, col_vw_return), MergedCell)
    ):
        excel_row += 1

    # Date
    ws.cell(excel_row, col_date, row['Date'])

    # MV Return
    ws.cell(excel_row, col_mv_return, row['MV_Return'])

    # VW Return
    ws.cell(excel_row, col_vw_return, row['VW_Return'])

print(f"   ✓ Filled {len(results)} months of data")

# ============================================================
# 6. FILL SUMMARY STATISTICS
# ============================================================

print("\n[6] Adding summary statistics...")

# Find good location for summary stats (to the right or below)
summary_start_col = 5  # Column E
summary_start_row = header_row

# Headers
ws.cell(summary_start_row, summary_start_col, "Metric")
ws.cell(summary_start_row, summary_start_col + 1, "Minimum Variance")
ws.cell(summary_start_row, summary_start_col + 2, "Value Weighted")

# Format
for col in range(summary_start_col, summary_start_col + 3):
    cell = ws.cell(summary_start_row, col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Write metrics
metrics = [
    'Annualized Return',
    'Annualized Volatility',
    'Sharpe Ratio',
    'Min Monthly Return',
    'Max Monthly Return',
    'Cumulative Return'
]

for i, metric in enumerate(metrics):
    row = summary_start_row + 1 + i
    
    # Metric name
    ws.cell(row, summary_start_col, metric)
    ws.cell(row, summary_start_col).font = Font(bold=True)
    
    # MV Portfolio
    mv_value = summary.loc['Minimum Variance', metric]
    ws.cell(row, summary_start_col + 1, mv_value)
    
    # VW Portfolio
    vw_value = summary.loc['Value Weighted', metric]
    ws.cell(row, summary_start_col + 2, vw_value)
    
    # Format percentages for relevant metrics
    if 'Return' in metric or 'Volatility' in metric:
        ws.cell(row, summary_start_col + 1).number_format = '0.00%'
        ws.cell(row, summary_start_col + 2).number_format = '0.00%'
    elif 'Ratio' in metric:
        ws.cell(row, summary_start_col + 1).number_format = '0.0000'
        ws.cell(row, summary_start_col + 2).number_format = '0.0000'

print(f"   ✓ Added {len(metrics)} metrics")

# ============================================================
# 7. FORMAT MONTHLY RETURNS
# ============================================================

print("\n[7] Formatting monthly returns...")

# Format return columns as percentages
for idx in range(len(results)):
    excel_row = data_start_row + idx
    ws.cell(excel_row, col_mv_return).number_format = '0.00%'
    ws.cell(excel_row, col_vw_return).number_format = '0.00%'

# Auto-adjust column widths
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 20
ws.column_dimensions['E'].width = 22
ws.column_dimensions['F'].width = 18
ws.column_dimensions['G'].width = 18

print("   ✓ Formatting applied")

# ============================================================
# 8. ADD METADATA
# ============================================================

print("\n[8] Adding metadata...")

metadata_row = data_start_row + len(results) + 2

ws.cell(metadata_row, 1, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
ws.cell(metadata_row + 1, 1, f"Region: Pacific Area")
ws.cell(metadata_row + 2, 1, f"Scope: Scope 1")
ws.cell(metadata_row + 3, 1, f"Period: Jan 2014 - Dec 2025")

for row in range(metadata_row, metadata_row + 4):
    ws.cell(row, 1).font = Font(italic=True, size=9, color="666666")

# ============================================================
# 9. SAVE FILLED TEMPLATE
# ============================================================

print("\n[9] Saving filled template...")

output_filename = 'Template-Part-I-FILLED.xlsx'
wb.save(output_filename)

print(f"   ✓ Saved: {output_filename}")

# ============================================================
# 10. SUMMARY
# ============================================================

print("\n" + "=" * 60)
print("TEMPLATE FILLING COMPLETE")
print("=" * 60)

print(f"\n✅ Output file: {output_filename}")
print(f"\n📊 Data filled:")
print(f"   • {len(results)} months of returns (Jan 2014 - Dec 2025)")
print(f"   • {len(metrics)} summary statistics")
print(f"   • Formatted as percentages and ratios")

print(f"\n📍 Data location in template:")
print(f"   • Monthly returns: Columns A-C, starting row {data_start_row}")
print(f"   • Summary stats: Columns E-G, starting row {summary_start_row}")

print("\n" + "=" * 60)
print("READY FOR SUBMISSION")
print("=" * 60)

print("\nNext steps:")
print("1. Open Template-Part-I-FILLED.xlsx")
print("2. Review the filled data")
print("3. Adjust formatting if needed")
print("4. Submit by April 12, 2026 at midnight")

wb.close()
