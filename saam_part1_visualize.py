# ============================================================
# SAAM PROJECT - PART I: VISUALIZATION & TEMPLATE FILLER
# Creates plots and fills Template for Part I-SAAM.xlsx
# ============================================================

import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

# Set style
sns.set_style("whitegrid")
plt.rcParams['figure.figsize'] = (12, 6)
plt.rcParams['font.size'] = 10

print("=" * 60)
print("PART I: VISUALIZATION & TEMPLATE FILLER")
print("=" * 60)

# directories
RESULTS_DIR = "resultsPart1"
OUTPUT_DIR = "part1_outputs"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ============================================================
# 1. LOAD RESULTS
# ============================================================

print("\n[1] Loading results...")

results_path = os.path.join(RESULTS_DIR, "part1_results.csv")
summary_path = os.path.join(RESULTS_DIR, "part1_summary_statistics.csv")

results = pd.read_csv(results_path)
results['Date'] = pd.to_datetime(results['Date'])

summary = pd.read_csv(summary_path, index_col=0)

print(f"   Results loaded from '{results_path}': {len(results)} months")
print(f"   Date range: {results['Date'].min()} to {results['Date'].max()}")

# ============================================================
# 2. CREATE CUMULATIVE RETURN PLOT
# ============================================================

print("\n[2] Creating cumulative return plot...")

fig, ax = plt.subplots(figsize=(14, 7))

# Plot cumulative returns
ax.plot(results['Date'], results['MV_CumReturn'], 
        label='Minimum Variance Portfolio', linewidth=2, color='#2E86AB')
ax.plot(results['Date'], results['VW_CumReturn'], 
        label='Value-Weighted Portfolio', linewidth=2, color='#A23B72')

# Formatting
ax.set_xlabel('Date', fontsize=12, fontweight='bold')
ax.set_ylabel('Cumulative Return (Base = 1)', fontsize=12, fontweight='bold')
ax.set_title('Portfolio Performance: Minimum Variance vs Value-Weighted\nPacific Area | 2014-2025', 
             fontsize=14, fontweight='bold', pad=20)
ax.legend(loc='upper left', fontsize=11, frameon=True, shadow=True)
ax.grid(True, alpha=0.3)

# Add annotations
final_mv = results['MV_CumReturn'].iloc[-1]
final_vw = results['VW_CumReturn'].iloc[-1]
ax.text(0.02, 0.98, 
        f'Final Cumulative Returns:\nMV: {(final_mv-1)*100:.2f}%\nVW: {(final_vw-1)*100:.2f}%',
        transform=ax.transAxes, fontsize=10,
        verticalalignment='top', bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.5))

plt.tight_layout()
cum_plot_path = os.path.join(OUTPUT_DIR, "cumulative_returns_plot.png")
plt.savefig(cum_plot_path, dpi=300, bbox_inches='tight')
print(f"   ✅ Saved: {cum_plot_path}")

# ============================================================
# 3. CREATE ADDITIONAL PLOTS
# ============================================================

print("\n[3] Creating additional plots...")

# Monthly returns distribution
fig, axes = plt.subplots(1, 2, figsize=(14, 5))

axes[0].hist(results['MV_Return'], bins=30, alpha=0.7, color='#2E86AB', edgecolor='black', label='MV')
axes[0].hist(results['VW_Return'], bins=30, alpha=0.7, color='#A23B72', edgecolor='black', label='VW')
axes[0].set_xlabel('Monthly Return', fontweight='bold')
axes[0].set_ylabel('Frequency', fontweight='bold')
axes[0].set_title('Distribution of Monthly Returns', fontweight='bold')
axes[0].legend()
axes[0].grid(True, alpha=0.3)

# Rolling 12-month volatility
results['MV_Rolling_Vol'] = results['MV_Return'].rolling(12).std() * np.sqrt(12)
results['VW_Rolling_Vol'] = results['VW_Return'].rolling(12).std() * np.sqrt(12)

axes[1].plot(results['Date'], results['MV_Rolling_Vol'], label='MV', color='#2E86AB', linewidth=2)
axes[1].plot(results['Date'], results['VW_Rolling_Vol'], label='VW', color='#A23B72', linewidth=2)
axes[1].set_xlabel('Date', fontweight='bold')
axes[1].set_ylabel('Annualized Volatility', fontweight='bold')
axes[1].set_title('Rolling 12-Month Volatility', fontweight='bold')
axes[1].legend()
axes[1].grid(True, alpha=0.3)

plt.tight_layout()
add_plot_path = os.path.join(OUTPUT_DIR, "additional_analysis_plots.png")
plt.savefig(add_plot_path, dpi=300, bbox_inches='tight')
print(f"   ✅ Saved: {add_plot_path}")

# ============================================================
# 4. SUMMARY TABLE FOR REPORT
# ============================================================

print("\n[4] Creating summary table...")

# Format summary statistics for report
summary_formatted = summary.copy()
summary_formatted['Annualized Return'] = summary_formatted['Annualized Return'].apply(lambda x: f"{x*100:.2f}%")
summary_formatted['Annualized Volatility'] = summary_formatted['Annualized Volatility'].apply(lambda x: f"{x*100:.2f}%")
summary_formatted['Sharpe Ratio'] = summary_formatted['Sharpe Ratio'].apply(lambda x: f"{x:.4f}")
summary_formatted['Min Monthly Return'] = summary_formatted['Min Monthly Return'].apply(lambda x: f"{x*100:.2f}%")
summary_formatted['Max Monthly Return'] = summary_formatted['Max Monthly Return'].apply(lambda x: f"{x*100:.2f}%")
summary_formatted['Cumulative Return'] = summary_formatted['Cumulative Return'].apply(lambda x: f"{x*100:.2f}%")

print("\n" + "=" * 60)
print("SUMMARY STATISTICS")
print("=" * 60)
print(summary_formatted.to_string())

summary_table_path = os.path.join(OUTPUT_DIR, "summary_table_formatted.csv")
summary_formatted.to_csv(summary_table_path)
print(f"\n✅ Saved: {summary_table_path}")

# ============================================================
# 5. FILL TEMPLATE (if template file is available)
# ============================================================

print("\n[5] Checking for template file...")

try:
    # Try to load template
    template_path = 'Template for Part I-SAAM.xlsx'
    wb = load_workbook(template_path)
    ws = wb.active
    
    print(f"   Template found: {template_path}")
    print("   Filling in data...")
    
    # Assuming template structure:
    # Column A: Date
    # Column B: MV Returns
    # Column C: VW Returns
    # Summary statistics section elsewhere
    
    # Write dates and returns (starting from row 2, skipping merged cells)
    for idx, row in results.iterrows():
        excel_row = idx + 2  # +2 because Excel is 1-indexed and row 1 is header

        # Skip down if current row overlaps merged cells in A–C
        while (
            isinstance(ws[f'A{excel_row}'], MergedCell)
            or isinstance(ws[f'B{excel_row}'], MergedCell)
            or isinstance(ws[f'C{excel_row}'], MergedCell)
        ):
            excel_row += 1

        ws[f'A{excel_row}'] = row['Date']
        ws[f'B{excel_row}'] = row['MV_Return']
        ws[f'C{excel_row}'] = row['VW_Return']
    
    # Write summary statistics (example locations - adjust as needed)
    # MV Portfolio summary at F2:F7
    # VW Portfolio summary at G2:G7
    
    summary_metrics = ['Annualized Return', 'Annualized Volatility', 'Sharpe Ratio', 
                      'Min Monthly Return', 'Max Monthly Return', 'Cumulative Return']
    
    for i, metric in enumerate(summary_metrics):
        ws[f'E{i+2}'] = metric
        ws[f'F{i+2}'] = summary.loc['Minimum Variance', metric]
        ws[f'G{i+2}'] = summary.loc['Value Weighted', metric]
    
    # Save filled template
    output_path = 'Template_Part1_FILLED.xlsx'
    wb.save(output_path)
    print(f"   ✅ Saved: {output_path}")
    
except FileNotFoundError:
    print("   ⚠️  Template file not found. Please fill manually using:")
    print("      - part1_results.csv (monthly returns)")
    print("      - part1_summary_statistics.csv (summary stats)")

# ============================================================
# 6. CREATE PORTFOLIO COMPOSITION SUMMARY
# ============================================================

print("\n[6] Analyzing portfolio compositions...")

comp_path = os.path.join(RESULTS_DIR, "part1_portfolio_compositions.csv")
comp = pd.read_csv(comp_path)

# Top holdings by year
print("\n" + "=" * 60)
print("TOP 10 HOLDINGS BY YEAR (Minimum Variance Portfolio)")
print("=" * 60)

for year in sorted(comp['Year'].unique()):
    year_data = comp[comp['Year'] == year].nlargest(10, 'Weight')
    print(f"\n{year}:")
    print(year_data[['ISIN', 'Weight']].to_string(index=False))

# Average concentration
print("\n" + "=" * 60)
print("PORTFOLIO CONCENTRATION METRICS")
print("=" * 60)

for year in sorted(comp['Year'].unique()):
    year_data = comp[comp['Year'] == year]
    n_firms = len(year_data)
    max_weight = year_data['Weight'].max()
    top5_weight = year_data.nlargest(5, 'Weight')['Weight'].sum()
    herfindahl = (year_data['Weight'] ** 2).sum()
    
    print(f"\n{year}:")
    print(f"   Number of firms: {n_firms}")
    print(f"   Max weight: {max_weight*100:.2f}%")
    print(f"   Top 5 weight: {top5_weight*100:.2f}%")
    print(f"   Herfindahl index: {herfindahl:.4f}")

# ============================================================
# 7. FINAL SUMMARY
# ============================================================

print("\n" + "=" * 60)
print("FILES GENERATED")
print("=" * 60)
print(f"\n✅ {cum_plot_path}")
print(f"✅ {add_plot_path}")
print(f"✅ {summary_table_path}")
print(f"✅ {results_path}")
print(f"✅ {summary_path}")
print(f"✅ {comp_path}")

print("\n" + "=" * 60)
print("PART I COMPLETE - READY FOR SUBMISSION")
print("=" * 60)
